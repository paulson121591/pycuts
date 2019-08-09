import sys
import random
from PySide2 import *
import pickle
import pyautogui
import PyQt5
from openpyxl import load_workbook
import os
from os import listdir
import shutil
import win32com.client
from win32com.client import Dispatch, constants
import re
import win32com.client



    

class MyWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        

        

        self.buttonLoad = QtWidgets.QPushButton("Create Production Email")
        self.buttonxQuote = QtWidgets.QPushButton("Create Quote Email")
        self.buttonNew = QtWidgets.QPushButton("Auto Collect Data")
        self.buttonSave = QtWidgets.QPushButton("Save Data")
        self.buttonxl = QtWidgets.QPushButton("Consolidate")
        self.buttonxApproval = QtWidgets.QPushButton("Create Approval")
        self.buttonxOrder = QtWidgets.QPushButton("Create Order Form")
        
        self.lJobNumber = QtWidgets.QLabel('Job Number')
        self.inputJobNumber = QtWidgets.QLineEdit()
        self.lJobname = QtWidgets.QLabel('Job Name')
        self.inputJobName = QtWidgets.QLineEdit()
        self.lSalesman = QtWidgets.QLabel('Salesman')
        self.inputSalesman = QtWidgets.QLineEdit()
        self.lDesigner = QtWidgets.QLabel('Designer')
        self.inputDesigner = QtWidgets.QLineEdit()
        self.lRegion = QtWidgets.QLabel('Region')
        self.inputRegion = QtWidgets.QLineEdit()
        self.lStreetName = QtWidgets.QLabel('Street Name')
        self.inputStreetName = QtWidgets.QLineEdit()
        self.lZipCode = QtWidgets.QLabel('Zip Code')
        self.inputZipCode = QtWidgets.QLineEdit()
        self.lQuotedPrice = QtWidgets.QLabel('Quoted Price')
        self.inputQuotedPrice = QtWidgets.QLineEdit()
        self.lTotalPrice = QtWidgets.QLabel('Total Price')
        self.inputTotalPrice = QtWidgets.QLineEdit()
        self.lCustomerCode = QtWidgets.QLabel('Customer Code')
        self.inputCustomerCode = QtWidgets.QLineEdit()
        self.lCustomerName = QtWidgets.QLabel('Customer Name')
        self.inputCustomerName = QtWidgets.QLineEdit()
        self.lBillingStreet = QtWidgets.QLabel('Billing Street')
        self.inputBillingStreet = QtWidgets.QLineEdit()
        self.lBillingCity = QtWidgets.QLabel('Billing City')
        self.inputBillingCity = QtWidgets.QLineEdit()
        self.lBillingZip = QtWidgets.QLabel('Billing Zip')
        self.inputBillingZip = QtWidgets.QLineEdit()
        self.lBf = QtWidgets.QLabel('BF')
        self.lPriceBf = QtWidgets.QLabel()
        self.inputBf = QtWidgets.QLineEdit()
        
        self.lPhoneNumber = QtWidgets.QLabel('Phone Number')
        self.inputPhoneNumber = QtWidgets.QLineEdit()
        self.lemail = QtWidgets.QLabel('Email')
        self.inputEmail = QtWidgets.QLineEdit()
        self.lDate = QtWidgets.QLabel('Delivery Date')
        self.inputDate = QtWidgets.QLineEdit()

                                
        
        
        
        


        priceBf=0.00



        self.layout = QtWidgets.QGridLayout()
        self.layout.addWidget(self.lJobNumber,1,0)
        self.layout.addWidget(self.inputJobNumber,1,1)
        self.layout.addWidget(self.lJobname,2,0)
        self.layout.addWidget(self.inputJobName,2,1)
        self.layout.addWidget(self.lSalesman,3,0)
        self.layout.addWidget(self.inputSalesman,3,1)
        self.layout.addWidget(self.lDesigner,4,0)
        self.layout.addWidget(self.inputDesigner,4,1)
        self.layout.addWidget(self.lRegion,5,0)
        self.layout.addWidget(self.inputRegion,5,1)
        self.layout.addWidget(self.lStreetName,6,0)
        self.layout.addWidget(self.inputStreetName,6,1)
        self.layout.addWidget(self.lZipCode,7,0)
        self.layout.addWidget(self.inputZipCode,7,1) 
        self.layout.addWidget(self.lQuotedPrice,8,0)
        self.layout.addWidget(self.inputQuotedPrice,8,1)
        self.layout.addWidget(self.lTotalPrice,9,0)        
        self.layout.addWidget(self.inputTotalPrice,9,1)
        self.layout.addWidget(self.lCustomerCode,10,0)         
        self.layout.addWidget(self.inputCustomerCode,10,1)
        self.layout.addWidget(self.lCustomerName,11,0)        
        self.layout.addWidget(self.inputCustomerName,11,1) 
        self.layout.addWidget(self.lBillingStreet,12,0)        
        self.layout.addWidget(self.inputBillingStreet,12,1) 
        self.layout.addWidget(self.lBillingCity,13,0)        
        self.layout.addWidget(self.inputBillingCity,13,1) 
        self.layout.addWidget(self.lBillingZip,14,0)        
        self.layout.addWidget(self.inputBillingZip,14,1)  
        self.layout.addWidget(self.lBf,15,0)  
        self.layout.addWidget(self.inputBf,15,1)
        self.layout.addWidget(self.lPriceBf,15,3)  
        self.layout.addWidget(self.lPhoneNumber,16,0)  
        self.layout.addWidget(self.inputPhoneNumber,16,1)  
        self.layout.addWidget(self.lemail,17,0)  
        self.layout.addWidget(self.inputEmail,17,1)
        self.layout.addWidget(self.lDate,18,0)  
        self.layout.addWidget(self.inputDate,18,1)                
                                            
        self.layout.addWidget(self.buttonLoad,1,3)
        self.layout.addWidget(self.buttonNew,2,3)
        self.layout.addWidget(self.buttonSave,3,3)
        self.layout.addWidget(self.buttonxl,4,3)
        self.layout.addWidget(self.buttonxApproval,5,3)
        self.layout.addWidget(self.buttonxOrder,6,3)
        self.layout.addWidget(self.buttonxQuote,7,3)
    
        self.setLayout(self.layout)


        self.buttonLoad.clicked.connect(self.sendEmail)
        self.buttonNew.clicked.connect(self.new)
        self.buttonSave.clicked.connect(self.save)
        self.buttonxl.clicked.connect(self.xl)
        self.buttonxApproval.clicked.connect(self.approval)
        self.buttonxOrder.clicked.connect(self.order)
        self.inputJobNumber.textChanged[str].connect(self.load)
        self.buttonxQuote.clicked.connect(self.sendQuote)
        self.inputBf.textChanged[str].connect(self.calcBf)
        self.inputQuotedPrice.textChanged[str].connect(self.calcBf)


    def load(self):

        
        
        
        jobNumber = self.inputJobNumber.text()
        
    
        try:
            jobInfo = pickle.load( open( jobNumber, "rb" ))
            jobName = jobInfo.get('Job Name','')
            salesman = jobInfo.get('Salesman','')
            designer = jobInfo.get('Designer','')
            region = jobInfo.get('Region','')
            streetName = jobInfo.get('Street Name','')
            zipCode = jobInfo.get('Zip Code','')
            quotedPrice = jobInfo.get('Quoted Price','')
            totalPrice = jobInfo.get('Total Price','')
            customerCode = jobInfo.get('Customer Code','')
            customerName = jobInfo.get('Customer Name','')
            billingStreet = jobInfo.get('Billing Street','')
            billingCity = jobInfo.get('Billing City','')
            billingZip = jobInfo.get('Billing Zip','')
            totalBf = jobInfo.get('Bf','')
            phoneNumber = jobInfo.get('phonenumber','')
            email = jobInfo.get('email','')
            date = jobInfo.get('date','')
            
        
        
            
        
        
            self.inputJobNumber.setText(jobNumber)
            self.inputJobName.setText(jobName) 
            self.inputSalesman.setText(salesman)
            self.inputDesigner.setText(designer)
            self.inputRegion.setText(region)
            self.inputStreetName.setText(streetName)
            self.inputZipCode.setText(zipCode)
            self.inputQuotedPrice.setText(quotedPrice)
            self.inputTotalPrice.setText(totalPrice)
            self.inputCustomerCode.setText(customerCode)
            self.inputCustomerName.setText(customerName)
            self.inputBillingStreet.setText(billingStreet)
            self.inputBillingCity.setText(billingCity)
            self.inputBillingZip.setText(billingZip)
            self.inputBf.setText(totalBf)
            self.inputPhoneNumber.setText(phoneNumber)
            self.inputEmail.setText(email)
            self.inputDate.setText(date)
            
    
    
        except FileNotFoundError:
            
            self.inputJobName.setText('No Job Found') 
            self.inputSalesman.setText('')
            self.inputDesigner.setText('')
            self.inputRegion.setText('')
            self.inputStreetName.setText('')
            self.inputZipCode.setText('')
            self.inputQuotedPrice.setText('')
            self.inputTotalPrice.setText('')
            self.inputCustomerCode.setText('')
            self.inputCustomerName.setText('')
            self.inputBillingStreet.setText('')
            self.inputBillingCity.setText('')
            self.inputBillingZip.setText('')
            self.inputBf.setText('')
            self.inputPhoneNumber.setText('')
            self.inputEmail.setText('')
            self.inputDate.setText('')
        
        

        
        
        

        
    
        
        
    def new(self):
        import jobdat
        jobdat.new()
        
    def save(self):
        jobNumber = self.inputJobNumber.text() 
        jobName = self.inputJobName.text() 
        salesman= self.inputSalesman.text()
        designer= self.inputDesigner.text()
        region= self.inputRegion.text()
        street= self.inputStreetName.text()
        zipCode= self.inputZipCode.text()
        quotedPrice= self.inputQuotedPrice.text()
        totalPrice = self.inputTotalPrice.text()
        customerCode=  self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        totalBf = self.inputBf.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
        
    #save copied info with pickle
        jobInfo = {'jobnumber':jobNumber, 'Job Name': jobName, 'Salesman': salesman,'Designer':designer,'Region':region,'Street Name':street,'Zip Code':zipCode,'Quoted Price':quotedPrice,'Total Price':totalPrice,'Customer Code':customerCode,'Customer Name':customerName,'Billing Street':billingStreet,'Billing City':billingCity,'Billing Zip':billingZip, 'Bf':totalBf,'phonenumber':phoneNumber, 'email':email,'date':date} 
        pickle.dump( jobInfo, open( jobNumber, "wb") )
        
        
    def xl(self):
        jobNumber = self.inputJobNumber.text() 
        jobName = self.inputJobName.text() 
        salesman= self.inputSalesman.text()
        designer= self.inputDesigner.text()
        region= self.inputRegion.text()
        street= self.inputStreetName.text()
        zipCode= self.inputZipCode.text()
        quotedPrice= self.inputQuotedPrice.text()
        totalPrice = self.inputTotalPrice.text()
        customerCode=  self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        totalBf = self.inputBf.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
                
        wb = load_workbook('walk.xlsx')
        ws = wb['Entry']
        ws['C4'] = jobNumber
        ws['D4'] = customerCode
        ws['B4'] = jobName
        ws['E4'] = salesman
        ws['F4'] = salesman
        ws['G4'] = phoneNumber
        ws['B16'] = date
        ws['H4'] = email
        #TODO add state
        ws['I4'] = billingStreet+", "+billingCity+', '+billingZip

        ws['B10'] = totalBf
        ws['C10'] = quotedPrice
        ws['F10'] = float(zipCode)
        ws['C16'] = street
        ws['B19'] = customerName
        ws['C19'] = phoneNumber
        wb.save(jobNumber +'.xlsx')
        pyautogui.alert('Data Consolidate')
        
    def approval(self):
        
        try:
        
            jobNumber = self.inputJobNumber.text() 
            
            o = win32com.client.Dispatch("Excel.Application")

            o.Visible = False

            wb_path = os.getcwd()+'\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open(wb_path)



            ws_index_list = [3] #say you want to print these sheets

            path_to_pdf = os.getcwd()+'\\' + jobNumber + ' Approval.pdf'

            print_area = 'A1:L50'



            for index in ws_index_list:

                #off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area



            wb.WorkSheets(ws_index_list).Select()

            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
            wb.Close(True)
            pyautogui.alert('Approval Created')
            
            
            if not os.path.exists(r'O:\Jobs\\'+jobNumber+'\Order'):
                os.mkdir(r'O:\Jobs\\'+jobNumber+'\Order')
        
            file = jobNumber+' Approval.pdf'
            shutil.move(os.getcwd()+'\\'+file,'O:\Jobs\\'+jobNumber+'\Order\\'+file)


    
            pyautogui.alert('did it work?')
        except:
            pyautogui.alert('Unable to creatE pdf Have you Consolidated this job yet?')
            


    def order(self):
        import win32com.client
        try:
        
            jobNumber = self.inputJobNumber.text() 
            
            o = win32com.client.Dispatch("Excel.Application")

            o.Visible = False

            wb_path = os.getcwd()+'\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open(wb_path)



            ws_index_list = [4] #say you want to print these sheets

            path_to_pdf = os.getcwd()+'\\' + jobNumber + ' Order.pdf'

            print_area = 'A1:J50'
            
            



            for index in ws_index_list:

                #off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area



            wb.WorkSheets(ws_index_list).Select()

            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
            wb.Close(True)
            pyautogui.alert('Order Form Created')
            if not os.path.exists(r'O:\Jobs\\'+jobNumber+'\Order'):
                os.mkdir(r'O:\Jobs\\'+jobNumber+'\Order')
            
            file = jobNumber+' Order.pdf'
            shutil.move(os.getcwd()+'\\'+file,'O:\Jobs\\'+jobNumber+'\Order\\'+file)


        
            pyautogui.alert('did it work?')
        except:
            pyautogui.alert('Unable to creatE pdf Have you Consolidated this job yet?')
            
                    
        
    def sendEmail(self):
        jobNumber = self.inputJobNumber.text() 
        jobName=self.inputJobName.text() 
        phoneNumber=self.inputPhoneNumber.text() 
        date = self.inputDate.text() 
        customerName = self.inputCustomerName
        sub= str(jobNumber)+'-'+str(jobName)+'-'+'Order'
        const=win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch("Outlook.Application")
        newMail = obj.CreateItem(olMailItem)
        newMail.Subject = sub
        newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>This one is ready for production <br><br> Thanks,<br><br>Delivery Date:"+date+"<br>Call before Delivery "+phoneNumber+"<br><br>Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = "tstrayer@paneltruss.com; ty@paneltruss.com; mlowe@paneltruss.com; dickie@paneltruss.com; amarsingill@paneltruss.com; dlawrence@paneltruss.com; akimsey@paneltruss.com"

        newMail.display()
        #newMail.Send()
    def sendQuote(self):
        
        jobNumber = self.inputJobNumber.text() 
        jobName=self.inputJobName.text() 
        phoneNumber=self.inputPhoneNumber.text() 
        emailAddress= self.inputEmail.text()
        date = self.inputDate.text() 
        customerName = self.inputCustomerName
        sub= str(jobNumber)+'-'+str(jobName)+'-'+' Panel Truss Quote'
        const=win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch("Outlook.Application")
        newMail = obj.CreateItem(olMailItem)
        newMail.Subject = sub
        newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>Attached to this email it the quote for the trusses you requested.<br><br>Let me know if you have any questions <br>Thanks,<br><br>Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = emailAddress

        newMail.display()     
    def calcBf(self):
        bf = float(self.inputBf.text())
        quotedPrice = self.inputQuotedPrice.text()
        quotedPrice = re.sub('[!@#$,]', '', quotedPrice)


        print (quotedPrice)
        quotedPrice = float(quotedPrice)
        bfPrice = quotedPrice/bf
        bfPrice = round(bfPrice,2)
        bfPrice = 'Price/BF: $'+str(bfPrice)
        self.lPriceBf.setText(bfPrice) 
           
        
                
        
        
if __name__ == "__main__":
    app = QtWidgets.QApplication([])

    widget = MyWidget()
    widget.resize(800, 600)
    widget.show()

    sys.exit(app.exec_())