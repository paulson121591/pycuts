import sys
import pickle
import pyautogui
import random
import os
import shutil
import win32com.client
import re


from win32com.client import Dispatch, constants

from os import listdir

from openpyxl import load_workbook

import PySide2 
from PySide2.QtUiTools import QUiLoader
from PySide2.QtWidgets import QApplication, QPushButton, QLineEdit, QLabel, QComboBox,QSpinBox,QRadioButton,QDoubleSpinBox,QCheckBox,QTableWidget,QTableWidgetItem
from PySide2.QtCore import QFile, QObject
 
class Form(QObject):
 #converting qt main file to python so python can edit and get information from it 
    def __init__(self, ui_file, parent=None):
        super(Form, self).__init__(parent)
        ui_file = QFile(ui_file)
        ui_file.open(QFile.ReadOnly)
 
        loader = QUiLoader()
        self.window = loader.load(ui_file)
        ui_file.close()
 
        self.inputJobNumber = self.window.findChild(QLineEdit, 'inputJobNumber')
        self.inputJobName = self.window.findChild(QLineEdit, 'inputJobName')
        self.inputSalesman = self.window.findChild(QLineEdit, 'inputSalesman')
        self.inputDesigner = self.window.findChild(QLineEdit, 'inputDesigner')
        self.inputRegion = self.window.findChild(QLineEdit, 'inputRegion')
        self.inputStreetName = self.window.findChild(QLineEdit, 'inputStreetName')
        self.inputZipCode = self.window.findChild(QLineEdit, 'inputZipCode')
        self.inputQuotedPrice = self.window.findChild(QLineEdit, 'inputQuotedPrice')
        self.inputTotalPrice = self.window.findChild(QLineEdit, 'inputTotalPrice')
        self.inputCustomerCode = self.window.findChild(QLineEdit, 'inputCustomerCode')
        self.inputCustomerName = self.window.findChild(QLineEdit, 'inputCustomerName')
        self.inputBillingStreet = self.window.findChild(QLineEdit, 'inputBillingStreet')
        self.inputBillingCity = self.window.findChild(QLineEdit, 'inputBillingCity')
        self.inputBillingZip = self.window.findChild(QLineEdit, 'inputBillingZip')
        self.inputBf = self.window.findChild(QLineEdit, 'inputBf')
        self.inputPhoneNumber = self.window.findChild(QLineEdit, 'inputPhoneNumber')
        self.inputEmail = self.window.findChild(QLineEdit, 'inputEmail')
        self.inputDate = self.window.findChild(QLineEdit, 'inputDate')
        self.lPriceBf = self.window.findChild(QLabel, 'lPriceBf')
        self.pitch = self.window.findChild(QComboBox, 'pitch')
        self.span = self.window.findChild(QSpinBox, 'span')
        self.gable = self.window.findChild(QSpinBox, 'gable')
        self.trussNumber = self.window.findChild(QLineEdit, 'trussNumber')
        self.pPerBf = self.window.findChild(QDoubleSpinBox, 'pPerBf')
        self.lPrice = self.window.findChild(QLabel, 'lPrice')
        self.lTotalBf = self.window.findChild(QLabel, 'lTotalBf')
        self.taxRate = self.window.findChild(QLabel, 'taxRate')
        self.taxFree = self.window.findChild(QCheckBox, 'taxFree')

        self.inputPo = self.window.findChild(QLineEdit, 'inputPo')
        self.tableJobEst = self.window.findChild(QTableWidget,'tableJobEst')
        self.jobNameEst = self.window.findChild(QLineEdit, 'jobNameEst')
        self.runNameEst = self.window.findChild(QLineEdit, 'runNameEst')
        self.bcSize = self.window.findChild(QComboBox, 'bcSize')
       
        self.depthFloor = self.window.findChild(QComboBox, 'depthFloor')
        self.spanFloor = self.window.findChild(QSpinBox, 'spanFloor')
        self.trussNumberFloor = self.window.findChild(QLineEdit, 'trussNumberFloor')
        
        
        

    
 
        buttonLoad = self.window.findChild(QPushButton, 'buttonLoad')
        buttonxQuote = self.window.findChild(QPushButton, 'buttonxQuote')
        buttonNew = self.window.findChild(QPushButton, 'buttonNew')
        buttonSave = self.window.findChild(QPushButton, 'buttonSave')
        buttonxl = self.window.findChild(QPushButton, 'buttonxl')
        buttonxApproval = self.window.findChild(QPushButton, 'buttonxApproval')
        buttonxQuoteForm = self.window.findChild(QPushButton, 'buttonxQuoteForm')
        buttonxOrder = self.window.findChild(QPushButton, 'buttonxOrder')
        buttoncalc = self.window.findChild(QPushButton, 'calculate')
        calculateFloor = self.window.findChild(QPushButton, 'calculateFloor')
        addToJobEst = self.window.findChild(QPushButton, 'addToJobEst')
        jobTotalEst = self.window.findChild(QPushButton, 'jobTotalEst')
        deleteEst = self.window.findChild(QPushButton, 'deleteEst')
        

        buttonLoad.clicked.connect(self.sendEmail)
        deleteEst.clicked.connect(self.deleteSel)
        buttonNew.clicked.connect(self.new)
        buttonSave.clicked.connect(self.save)
        buttonxl.clicked.connect(self.xl)
        buttonxApproval.clicked.connect(self.approval)
        buttonxOrder.clicked.connect(self.order)
        buttoncalc.clicked.connect(self.bfEst)
        calculateFloor.clicked.connect(self.bfEstFloor)
        addToJobEst.clicked.connect(self.jobManagment)
        jobTotalEst.clicked.connect(self.jobTotal)
        self.inputJobNumber.textChanged[str].connect(self.load)
        buttonxQuote.clicked.connect(self.sendQuote)
        self.inputBf.textChanged[str].connect(self.calcBf)
        self.inputQuotedPrice.textChanged[str].connect(self.calcBf)
        self.inputQuotedPrice.textChanged[str].connect(self.tax)
        
        self.inputZipCode.textChanged[str].connect(self.tax)
        self.jobNameEst.textChanged[str].connect(self.findJob)
        buttonxQuoteForm.clicked.connect(self.quoteForm)

        self.taxFree.stateChanged.connect(self.tax)
        
        
        self.window.show()
        
    
    
    
# loads the pickle file when you type a job Number in the job number input
    def load(self):
    
        
        
        
        jobNumber = self.inputJobNumber.text()
        
    
        try:
            jobInfo = pickle.load( open( jobNumber, "rb" ))
            jobName = jobInfo.get('Job Name','')
            salesman = jobInfo.get('Salesman','')
            designer = jobInfo.get('Designer','')
            region = jobInfo.get('Region','')
            streetName = jobInfo.get('Street Name','')
            zipCode = jobInfo.get('Zip Code','')
            quotedPrice = jobInfo.get('Quoted Price','')
            totalPrice = jobInfo.get('Total Price','')
            customerCode = jobInfo.get('Customer Code','')
            customerName = jobInfo.get('Customer Name','')
            billingStreet = jobInfo.get('Billing Street','')
            billingCity = jobInfo.get('Billing City','')
            billingZip = jobInfo.get('Billing Zip','')
            totalBf = jobInfo.get('Bf','')
            phoneNumber = jobInfo.get('phonenumber','')
            email = jobInfo.get('email','')
            date = jobInfo.get('date','')
            po = jobInfo.get('po','')
            
        
        
            
        
        
            self.inputJobNumber.setText(jobNumber)
            self.inputJobName.setText(jobName) 
            self.inputSalesman.setText(salesman)
            self.inputDesigner.setText(designer)
            self.inputRegion.setText(region)
            self.inputStreetName.setText(streetName)
            self.inputZipCode.setText(zipCode)
            self.inputQuotedPrice.setText(quotedPrice)
            self.inputTotalPrice.setText(totalPrice)
            self.inputCustomerCode.setText(customerCode)
            self.inputCustomerName.setText(customerName)
            self.inputBillingStreet.setText(billingStreet)
            self.inputBillingCity.setText(billingCity)
            self.inputBillingZip.setText(billingZip)
            self.inputBf.setText(totalBf)
            self.inputPhoneNumber.setText(phoneNumber)
            self.inputEmail.setText(email)
            self.inputDate.setText(date)
            self.inputPo.setText(po)
            
    
     #if no job found    
        except FileNotFoundError:
            
            self.inputJobName.setText('No Job Found') 
            self.inputSalesman.setText('')
            self.inputDesigner.setText('')
            self.inputRegion.setText('')
            self.inputStreetName.setText('')
            self.inputZipCode.setText('')
            self.inputQuotedPrice.setText('')
            self.inputTotalPrice.setText('')
            self.inputCustomerCode.setText('')
            self.inputCustomerName.setText('')
            self.inputBillingStreet.setText('')
            self.inputBillingCity.setText('')
            self.inputBillingZip.setText('')
            self.inputBf.setText('')
            self.inputPhoneNumber.setText('')
            self.inputEmail.setText('')
            self.inputDate.setText('')
        
        
#runs jobdat.py which gets the data from icommand
    def new(self):
        import jobdat
        jobdat.new()


#saves info to a pickle file       
    def save(self):
        jobNumber = self.inputJobNumber.text() 
        jobName = self.inputJobName.text() 
        salesman= self.inputSalesman.text()
        designer= self.inputDesigner.text()
        region= self.inputRegion.text()
        street= self.inputStreetName.text()
        zipCode= self.inputZipCode.text()
        quotedPrice= self.inputQuotedPrice.text()
        totalPrice = self.inputTotalPrice.text()
        customerCode=  self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        totalBf = self.inputBf.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
        po = self.inputPo.text()
        
    
        jobInfo = {'jobnumber':jobNumber, 'Job Name': jobName, 'Salesman': salesman,'Designer':designer,'Region':region,'Street Name':street,'Zip Code':zipCode,'Quoted Price':quotedPrice,'Total Price':totalPrice,'Customer Code':customerCode,'Customer Name':customerName,'Billing Street':billingStreet,'Billing City':billingCity,'Billing Zip':billingZip, 'Bf':totalBf,'phonenumber':phoneNumber, 'email':email,'date':date,'po':po} 
        pickle.dump( jobInfo, open( jobNumber, "wb") )
        
#creates and xl file to print pdfs from       
    def xl(self):
        jobNumber = self.inputJobNumber.text() 
        jobName = self.inputJobName.text() 
        salesman= self.inputSalesman.text()
        designer= self.inputDesigner.text()
        region= self.inputRegion.text()
        street= self.inputStreetName.text()
        zipCode= self.inputZipCode.text()
        quotedPrice= self.inputQuotedPrice.text()
        totalPrice = self.inputTotalPrice.text()
        customerCode=  self.inputCustomerCode.text()
        customerName = self.inputCustomerName.text()
        billingStreet = self.inputBillingStreet.text()
        billingCity = self.inputBillingCity.text()
        billingZip = self.inputBillingZip.text()
        totalBf = self.inputBf.text()
        phoneNumber = self.inputPhoneNumber.text()
        email = self.inputEmail.text()
        date = self.inputDate.text()
        po = self.inputPo.text()
        
                
        wb = load_workbook('walk.xlsx')
        ws = wb['Entry']
        ws['C4'] = jobNumber
        ws['D4'] = customerCode
        ws['B4'] = jobName
        ws['E4'] = salesman
        ws['F4'] = salesman
        ws['G4'] = phoneNumber
        ws['B16'] = date
        ws['H4'] = email
        #TODO add state
        ws['I4'] = billingStreet+", "+billingCity+', '+billingZip
        if self.taxFree.isChecked():
            ws['J10']=True
            pyautogui.alert('Warning: No Tax applied ')
            
        

        ws['B10'] = totalBf
        ws['C10'] = quotedPrice
        ws['F10'] = float(zipCode)
        ws['C16'] = street
        ws['B19'] = customerName
        ws['C19'] = phoneNumber
        ws['F16']
        wb.save(jobNumber +'.xlsx')
        pyautogui.alert('Data Consolidate')
        
        
#Creates pdf approval form and puts it in the job folder 
    def approval(self):
        try:
            import win32com.client
        
            jobNumber = self.inputJobNumber.text() 
            
            
            o = win32com.client.Dispatch("Excel.Application")
            
            

            o.Visible = True
            
            

            wb_path = os.getcwd()+'\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open(wb_path)
            



            ws_index_list = [3] #say you want to print these sheets

            path_to_pdf = os.getcwd()+'\\' + jobNumber + ' Approval.pdf'

            print_area = 'A1:L50'



            for index in ws_index_list:

                #off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area



            wb.WorkSheets(ws_index_list).Select()

            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
            wb.Close(True)
            pyautogui.alert('Approval Created')
            
            
            if not os.path.exists(r'O:\Jobs\\'+jobNumber+'\Order'):
                os.mkdir(r'O:\Jobs\\'+jobNumber+'\Order')
        
            file = jobNumber+' Approval.pdf'
            shutil.move(os.getcwd()+'\\'+file,'O:\Jobs\\'+jobNumber+'\Order\\'+file)


    
            
        except:
            pyautogui.alert('Unable to creatE pdf Have you Consolidated this job yet?')
            

#Creates pdf approval form and puts it in the job folder 
    def order(self):
        
        try:
        
            jobNumber = self.inputJobNumber.text() 
            
            o = win32com.client.Dispatch("Excel.Application")

            o.Visible = False

            wb_path = os.getcwd()+'\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open(wb_path)



            ws_index_list = [4] #say you want to print these sheets

            path_to_pdf = os.getcwd()+'\\' + jobNumber + ' Order.pdf'

            print_area = 'A1:J50'
            
            



            for index in ws_index_list:

                #off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area



            wb.WorkSheets(ws_index_list).Select()

            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
            wb.Close(True)
            pyautogui.alert('Order Form Created')
            if not os.path.exists(r'O:\Jobs\\'+jobNumber+'\Order'):
                os.mkdir(r'O:\Jobs\\'+jobNumber+'\Order')
            
            file = jobNumber+' Order.pdf'
            shutil.move(os.getcwd()+'\\'+file,'O:\Jobs\\'+jobNumber+'\Order\\'+file)


        
            pyautogui.alert('did it work?')
        except:
            pyautogui.alert('Unable to creatE pdf Have you Consolidated this job yet?')
            
                    
#Creates email to send to Production using outlook    
    def sendEmail(self):
        jobNumber = self.inputJobNumber.text() 
        jobName=self.inputJobName.text() 
        phoneNumber=self.inputPhoneNumber.text() 
        date = self.inputDate.text() 
        customerName = self.inputCustomerName
        sub= str(jobNumber)+'-'+str(jobName)+'-'+'Order'
        const=win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch("Outlook.Application")
        newMail = obj.CreateItem(olMailItem)
        newMail.Subject = sub
        newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>This one is ready for production <br><br> Thanks,<br><br>Delivery Date:"+date+"<br>Call before Delivery "+phoneNumber+"<br><br>Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = "tstrayer@paneltruss.com; ty@paneltruss.com; mlowe@paneltruss.com; dickie@paneltruss.com; amarsingill@paneltruss.com; dlawrence@paneltruss.com; akimsey@paneltruss.com"

        newMail.display()
        #newMail.Send()

#Creates email to send to customer using outlook 
    def sendQuote(self):
        
        jobNumber = self.inputJobNumber.text() 
        jobName=self.inputJobName.text() 
        phoneNumber=self.inputPhoneNumber.text() 
        emailAddress= self.inputEmail.text()
        date = self.inputDate.text() 
        customerName = self.inputCustomerName
        sub= str(jobNumber)+'-'+str(jobName)+'-'+' Panel Truss Quote'
        const=win32com.client.constants
        olMailItem = 0x0
        obj = win32com.client.Dispatch("Outlook.Application")
        newMail = obj.CreateItem(olMailItem)
        newMail.Subject = sub
        newMail.BodyFormat = 2 # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
        newMail.HTMLBody = "<HTML><BODY>Attached to this email it the quote for the trusses you requested.<br><br>Let me know if you have any questions <br>Thanks,<br><br>Paul Sfalanga III<br>(864)772-3423</BODY></HTML>"
        newMail.To = emailAddress

        newMail.display()     
   
#fills the price pre bf label with the correct price/ BF
    def calcBf(self):
        bf = float(self.inputBf.text())
        quotedPrice = self.inputQuotedPrice.text()
        quotedPrice = re.sub('[!@#$,]', '', quotedPrice)


        print (quotedPrice)
        quotedPrice = float(quotedPrice)
        bfPrice = quotedPrice/bf
        bfPrice = round(bfPrice,2)
        bfPrice = 'Price/BF: $'+str(bfPrice)
        self.lPriceBf.setText(bfPrice) 
        
        
 

    def bfEst(self):
        
        
        pitch = str(self.pitch.currentText())
        span = str(self.span.value())
        gable = str(self.gable.value())
        trussNumber = self.trussNumber.text()
        pPerBf = self.pPerBf.value()
        bfCode = 'p'+ pitch + 's'+ span
        
        
        if self.bcSize.currentText() == '2x4':
            bfCalcDict = {
                'p3/12s10': 15.3333,     
                'p3/12s11': 17.3333,  
                'p3/12s12': 18.6667,  
                'p3/12s13': 20,  
                'p3/12s14': 21.3333,  
                'p3/12s15': 22.6667,  
                'p3/12s16': 25.3333,  
                'p3/12s17': 30.64,  
                'p3/12s18': 33.3333, 
                'p3/12s19': 34.6667,  
                'p3/12s20': 37.3333,  
                'p3/12s21': 40,  
                'p3/12s22': 41.3333,  
                'p3/12s23': 42.6667,  
                'p3/12s24': 45.3333,  
                'p3/12s25': 48.6667,  
                'p3/12s26': 48.6667,  
                'p3/12s27': 52,  
                'p3/12s28': 54.6667,  
                'p3/12s29': 56,  
                'p3/12s30': 56,  
                'p3/12s31': 60.6667,  
                'p3/12s32': 60,  
                'p3/12s33': 68,  
                'p3/12s34': 70.6667,  
                'p3/12s35': 74.6667,  
                'p3/12s36': 86.6667,  
                'p3/12s37': 88.6667,  
                'p3/12s38': 88.6667,  
                'p3/12s39': 100.6667,  
                'p3/12s40': 99.3333,  
                'p4/12s10': 16,  
                'p4/12s11': 17.3333,  
                'p4/12s12': 18.6667,  
                'p4/12s13': 20,  
                'p4/12s14': 22,  
                'p4/12s15': 26,  
                'p4/12s16': 26,  
                'p4/12s17': 33.3333,  
                'p4/12s18': 34.6667,  
                'p4/12s19': 40,  
                'p4/12s20': 40,  
                'p4/12s21': 41.3333,  
                'p4/12s22': 41.3333,  
                'p4/12s23': 46.6667,  
                'p4/12s24': 46.6667,  
                'p4/12s25': 49.3333,  
                'p4/12s26': 52.6667,  
                'p4/12s27': 56.6667,  
                'p4/12s28': 56.6667, 
                'p4/12s29': 58,  
                'p4/12s30': 59.3333,  
                'p4/12s31':61.3333,
                'p4/12s32':69.3333,
                'p4/12s33': 74.6667,  
                'p4/12s34': 77.3333,  
                'p4/12s35': 78.6667,  
                'p4/12s36': 78.6667,  
                'p4/12s37': 81.3333,  
                'p4/12s38': 85.3333,  
                'p4/12s39': 86.6667, 
                'p4/12s40': 91.3333, 
                'p5/12s10': 16,  
                'p5/12s11': 19.3333,  
                'p5/12s12': 19.3333,  
                'p5/12s13': 22,  
                'p5/12s14': 22,  
                'p5/12s15': 26,  
                'p5/12s16': 26.6667, 
                'p5/12s17': 36,  
                'p5/12s18': 36,  
                'p5/12s19': 40,  
                'p5/12s20': 40,  
                'p5/12s21': 42.6667,  
                'p5/12s22': 45.3333,  
                'p5/12s23': 46.6667,  
                'p5/12s24': 50,  
                'p5/12s25': 52.6667,  
                'p5/12s26': 56,  
                'p5/12s27': 57.3333,  
                'p5/12s28': 57.3333, 
                'p5/12s29': 60,  
                'p5/12s30': 62.6667,  
                'p5/12s31':64,
                'p5/12s32': 74.6667,  
                'p5/12s33': 77.3333,  
                'p5/12s34': 81.3333,  
                'p5/12s35':82.6667,
                'p5/12s36': 85.3333,  
                'p5/12s37':89.3333,
                'p5/12s38':90.6667,
                'p5/12s39': 93.3333, 
                'p5/12s40': 97.3333,  
                'p6/12s10': 16.6667,  
                'p6/12s11': 19.3333,  
                'p6/12s12': 19.3333,  
                'p6/12s13': 22,  
                'p6/12s14': 22.6667,  
                'p6/12s15': 26.6667, 
                'p6/12s16': 26.6667, 
                'p6/12s17': 36,  
                'p6/12s18': 40,  
                'p6/12s19': 41.3333,  
                'p6/12s20': 41.3333,  
                'p6/12s21': 44,  
                'p6/12s22': 48,  
                'p6/12s23': 49.3333,  
                'p6/12s24': 52,  
                'p6/12s25': 57.3333,  
                'p6/12s26': 58,  
                'p6/12s27': 60.6667,  
                'p6/12s28': 60.6667,  
                'p6/12s29': 63.3333,  
                'p6/12s30': 64,  
                'p6/12s31': 70,  
                'p6/12s32': 82.6667,  
                'p6/12s33': 84,  
                'p6/12s34':85.3333,
                'p6/12s35': 89.3333,  
                'p6/12s36': 92,  
                'p6/12s37': 93.3333,  
                'p6/12s38': 93.3333,  
                'p6/12s39': 100,  
                'p6/12s40': 104,  
                'p7/12s10': 16.6667,  
                'p7/12s11': 19.3333,  
                'p7/12s12': 21.3333,  
                'p7/12s13': 22.6667,  
                'p7/12s14': 25.3333,  
                'p7/12s15': 27.3333,  
                'p7/12s16': 36,  
                'p7/12s17': 40, 
                'p7/12s18': 40,  
                'p7/12s19': 44,  
                'p7/12s20': 44,  
                'p7/12s21': 48,  
                'p7/12s22': 49.3333,  
                'p7/12s23': 50.6667,  
                'p7/12s24': 56.6667,  
                'p7/12s25': 58.6667,  
                'p7/12s26':60,
                'p7/12s27':61.3333,
                'p7/12s28': 64,  
                'p7/12s29': 66.6667,  
                'p7/12s30': 72,  
                'p7/12s31': 86.6667,  
                'p7/12s32': 86.6667,  
                'p7/12s33': 88,  
                'p7/12s34': 90.6667,  
                'p7/12s35': 98.6667,  
                'p7/12s36': 98.6667,  
                'p7/12s37': 100,  
                'p7/12s38': 102.6667,  
                'p7/12s39': 106.6667,
                'p7/12s40': 106.6667,    
                'p8/12s10': 18.6667,  
                'p8/12s11': 20,  
                'p8/12s12': 21.3333,  
                'p8/12s13': 26,  
                'p8/12s14': 26, 
                'p8/12s15': 27.3333,  
                'p8/12s16': 36,  
                'p8/12s17': 42.6667,  
                'p8/12s18': 42.6667,  
                'p8/12s19': 44,  
                'p8/12s20': 48,  
                'p8/12s21': 49.3333,  
                'p8/12s22': 49.3333,  
                'p8/12s23': 57.3333,  
                'p8/12s24': 57.3333, 
                'p8/12s25': 62.6667,  
                'p8/12s26':62.6667,
                'p8/12s27': 65.3333,  
                'p8/12s28': 68,  
                'p8/12s29': 70.6667,  
                'p8/12s30': 73.3333,  
                'p8/12s31':90.6667,
                'p8/12s32':90.6667,
                'p8/12s33': 97.3333,  
                'p8/12s34': 97.3333,  
                'p8/12s35': 102.6667,  
                'p8/12s36': 117.3333,  
                'p8/12s37': 121.3333,  
                'p8/12s38': 124,  
                'p8/12s39': 126.6667,  
                'p8/12s40': 130.6667,  
                'p9/12s10': 18.6667,  
                'p9/12s11': 21.3333,  
                'p9/12s12': 22,  
                'p9/12s13': 26, 
                'p9/12s14': 26, 
                'p9/12s15': 28,  
                'p9/12s16': 41.3333,  
                'p9/12s17': 42.6667, 
                'p9/12s18': 44,  
                'p9/12s19': 48,  
                'p9/12s20': 48,  
                'p9/12s21': 53.3333,  
                'p9/12s22': 53.3333,  
                'p9/12s23': 57.3333,  
                'p9/12s24': 61.3333, 
                'p9/12s25': 62.6667,  
                'p9/12s26': 64,  
                'p9/12s27':72,
                'p9/12s28': 72,  
                'p9/12s29': 76, 
                'p9/12s30': 76,  
                'p9/12s31': 92,  
                'p9/12s32': 102.6667,  
                'p9/12s33': 116,  
                'p9/12s34': 118.6667,  
                'p9/12s35': 122.6667,  
                'p9/12s36': 125.3334,  
                'p9/12s37': 129.3334,  
                'p9/12s38': 130.6666,  
                'p9/12s39': 150,  
                'p9/12s40': 154.6666,  
                'p10/12s10':18.6667,  
                'p10/12s11':22,  
                'p10/12s12':24.6667,  
                'p10/12s13':26.6667, 
                'p10/12s14':26.6667, 
                'p10/12s15':30.6667,  
                'p10/12s16':41.3333,  
                'p10/12s17':44, 
                'p10/12s18':44,  
                'p10/12s19':52,  
                'p10/12s20':52,  
                'p10/12s21':53.3333,  
                'p10/12s22':56, 
                'p10/12s23':61.3333,  
                'p10/12s24':61.3333,  
                'p10/12s25':66.6667,  
                'p10/12s26':72,  
                'p10/12s27':72,  
                'p10/12s28':74.6667,  
                'p10/12s29':141.3333,  
                'p10/12s30':98,   
                'p10/12s31':100.6667,   
                'p10/12s32':103.3334,   
                'p10/12s33':107.3333,   
                'p10/12s34':110,  
                'p10/12s35':114,   
                'p10/12s36':130.6667,   
                'p10/12s37':140,   
                'p10/12s38':142.6667,   
                'p10/12s39':145.3333,   
                'p10/12s40':148.6667,  
                'p11/12s10':20.6667,  
                'p11/12s11':22,  
                'p11/12s12':25.3333,  
                'p11/12s13':26.6667, 
                'p11/12s14':27.3333,  
                'p11/12s15':31.3333,  
                'p11/12s16':42.6667,  
                'p11/12s17':48, 
                'p11/12s18':50.6667,  
                'p11/12s19':52,  
                'p11/12s20':52,  
                'p11/12s21':57.3333,  
                'p11/12s22':60,  
                'p11/12s23':61.3333,  
                'p11/12s24':65.3333,  
                'p11/12s25':72, 
                'p11/12s26':70.6667, 
                'p11/12s27':92,   
                'p11/12s28':94.6667,   
                'p11/12s29':98.6667,   
                'p11/12s30':102.6667,   
                'p11/12s31':106.6667,   
                'p11/12s32':109.3334,   
                'p11/12s33':126.6666,   
                'p11/12s34':132.6666,   
                'p11/12s35':139.3334,   
                'p11/12s36':140,   
                'p11/12s37':144,   
                'p11/12s38':150,  
                'p11/12s39':152.6667,   
                'p11/12s40':154.6667,   
                'p12/12s10':20.6667,  
                'p12/12s11':25.3333,  
                'p12/12s12':25.3333,  
                'p12/12s13':27.3333,  
                'p12/12s14':30, 
                'p12/12s15':32,  
                'p12/12s16':46.6667,  
                'p12/12s17':50.6667,  
                'p12/12s18':50.6667,  
                'p12/12s19':52,  
                'p12/12s20':58.6667, 
                'p12/12s21':60,  
                'p12/12s22':60,  
                'p12/12s23':62.6667,  
                'p12/12s24':72,  
                'p12/12s25':86,  
                'p12/12s26':88.6667, 
                'p12/12s27':92.6667,   
                'p12/12s28':96.6667,   
                'p12/12s29':99.3334,   
                'p12/12s30':103.3334,  
                'p12/12s31':121.3334,  
                'p12/12s32':126.6666,   
                'p12/12s33':134,   
                'p12/12s34':136.6667,   
                'p12/12s35':142,   
                'p12/12s36':144,   
                'p12/12s37':147.3333,   
                'p12/12s38':150,   
                'p12/12s39':172.6667,   
                'p12/12s40':188.6667,   
                
                
            }
            
        if self.bcSize.currentText() == '2x6':
        
        
            bfCalcDict = {

                'p3/12s10' : 18,
                'p3/12s11' : 20,
                'p3/12s12' : 22.6667,
                'p3/12s13' : 24.6667,
                'p3/12s14' : 26,
                'p3/12s15' : 28,
                'p3/12s16' : 30.6667,
                'p3/12s17' : 38,
                'p3/12s18' : 39.3333,
                'p3/12s19' : 41.6667,
                'p3/12s20' : 44,
                'p3/12s21' : 46,
                'p3/12s22' : 47.3333,
                'p3/12s23' : 50.6667,
                'p3/12s24' : 53.3333,
                'p3/12s25' : 55.3333,
                'p3/12s26' : 55.3333,
                'p3/12s27' : 57.3333,
                'p3/12s28' : 64,
                'p3/12s29' : 66,
                'p3/12s30' : 66,
                'p3/12s31' : 69.6667,
                'p3/12s32' : 72,
                'p3/12s33' : 74,
                'p3/12s34' : 79.3333,
                'p3/12s35' : 84,
                'p3/12s36' : 86.6667,
                'p3/12s37' : 88.6667,
                'p3/12s38' : 88.6667,
                'p3/12s39' : 110,
                'p3/12s40' : 122,
                'p4/12s10' : 18,
                'p4/12s11' : 21.3333,
                'p4/12s12' : 22.6667,
                'p4/12s13' : 24.6667,
                'p4/12s14' : 26,
                'p4/12s15' : 31.3333,
                'p4/12s16' : 31.3333,
                'p4/12s17' : 39.3333,
                'p4/12s18' : 39.3333,
                'p4/12s19' : 45.3333,
                'p4/12s20' : 46.6667,
                'p4/12s21' : 48.6667,
                'p4/12s22' : 48.6667,
                'p4/12s23' : 53.3333,
                'p4/12s24' : 54.6667,
                'p4/12s25' : 56.6667,
                'p4/12s26' : 56.6667,
                'p4/12s27' : 66,
                'p4/12s28' : 66,
                'p4/12s29' : 68,
                'p4/12s30' : 68,
                'p4/12s31' : 73,
                'p4/12s32' : 72.6667,
                'p4/12s33' : 82,
                'p4/12s34' : 86,
                'p4/12s35' : 90.6667,
                'p4/12s36' : 90.6667,
                'p4/12s37' : 92.6667,
                'p4/12s38' : 96.6667,
                'p4/12s39' : 100,
                'p4/12s40' : 100,
                'p5/12s10' : 18,
                'p5/12s11' : 22.6667,
                'p5/12s12' : 23.3333,
                'p5/12s13' : 26.6667,
                'p5/12s14' : 26.6667,
                'p5/12s15' : 31.3333,
                'p5/12s16' : 31.3333,
                'p5/12s17' : 40.6667,
                'p5/12s18' : 42,
                'p5/12s19' : 46.6667,
                'p5/12s20' : 46.6667,
                'p5/12s21' : 50,
                'p5/12s22' : 52.6667,
                'p5/12s23' : 54.6667,
                'p5/12s24' : 54.6667,
                'p5/12s25' : 59.3333,
                'p5/12s26' : 62,
                'p5/12s27' : 66.6667,
                'p5/12s28' : 66.6667,
                'p5/12s29' : 70,
                'p5/12s30' : 71.3333,
                'p5/12s31' : 74.3333,
                'p5/12s32' : 78,
                'p5/12s33' : 85.3333,
                'p5/12s34' : 88,
                'p5/12s35' : 94.6667,
                'p5/12s36' : 94.6667,
                'p5/12s37' : 102,
                'p5/12s38' : 102,
                'p5/12s39' : 105.3333,
                'p5/12s40' : 105.3333,
                'p6/12s10' : 20,
                'p6/12s11' : 23.3333,
                'p6/12s12' : 23.3333,
                'p6/12s13' : 26.6667,
                'p6/12s14' : 27.3333,
                'p6/12s15' : 32,
                'p6/12s16' : 32,
                'p6/12s17' : 42,
                'p6/12s18' : 44.6667,
                'p6/12s19' : 48,
                'p6/12s20' : 48,
                'p6/12s21' : 50,
                'p6/12s22' : 55.3333,
                'p6/12s23' : 57.3333,
                'p6/12s24' : 57.3333,
                'p6/12s25' : 62,
                'p6/12s26' : 66.6667,
                'p6/12s27' : 68.6667,
                'p6/12s28' : 70,
                'p6/12s29' : 73.3333,
                'p6/12s30' : 74,
                'p6/12s31' : 77.6667,
                'p6/12s32' : 88,
                'p6/12s33' : 90,
                'p6/12s34' : 95.3333,
                'p6/12s35' : 98.6667,
                'p6/12s36' : 104,
                'p6/12s37' : 106,
                'p6/12s38' : 106,
                'p6/12s39' : 110.6667,
                'p6/12s40' : 116,
                'p7/12s10' : 20,
                'p7/12s11' : 23.3333,
                'p7/12s12' : 25.3333,
                'p7/12s13' : 27.3333,
                'p7/12s14' : 30,
                'p7/12s15' : 32,
                'p7/12s16' : 32.6667,
                'p7/12s17' : 46,
                'p7/12s18' : 46,
                'p7/12s19' : 48,
                'p7/12s20' : 50.6667,
                'p7/12s21' : 55.3333,
                'p7/12s22' : 55.3333,
                'p7/12s23' : 58.6667,
                'p7/12s24' : 61.3333,
                'p7/12s25' : 63.3333,
                'p7/12s26' : 67.3333,
                'p7/12s27' : 70.6667,
                'p7/12s28' : 72,
                'p7/12s29' : 76.6667,
                'p7/12s30' : 82,
                'p7/12s31' : 84.3333,
                'p7/12s32' : 90.6667,
                'p7/12s33' : 94,
                'p7/12s34' : 102,
                'p7/12s35' : 106.6667,
                'p7/12s36' : 110.6667,
                'p7/12s37' : 112.6667,
                'p7/12s38' : 115.3333,
                'p7/12s39' : 117.3333,
                'p7/12s40' : 117.3333,
                'p8/12s10' : 21.3333,
                'p8/12s11' : 24,
                'p8/12s12' : 25.3333,
                'p8/12s13' : 30,
                'p8/12s14' : 30.6667,
                'p8/12s15' : 32.6667,
                'p8/12s16' : 32.6667,
                'p8/12s17' : 46,
                'p8/12s18' : 48.6667,
                'p8/12s19' : 50.6667,
                'p8/12s20' : 53.3333,
                'p8/12s21' : 56.6667,
                'p8/12s22' : 56.6667,
                'p8/12s23' : 65.3333,
                'p8/12s24' : 65.3333,
                'p8/12s25' : 67.3333,
                'p8/12s26' : 71.3333,
                'p8/12s27' : 74.6667,
                'p8/12s28' : 77.3333,
                'p8/12s29' : 82,
                'p8/12s30' : 83.3333,
                'p8/12s31' : 85.6667,
                'p8/12s32' : 97.3333,
                'p8/12s33' : 102,
                'p8/12s34' : 108.6667,
                'p8/12s35' : 112,
                'p8/12s36' : 129.3333,
                'p8/12s37' : 135.3333,
                'p8/12s38' : 138,
                'p8/12s39' : 141.3333,
                'p8/12s40' : 145.3333,
                'p9/12s10' : 22,
                'p9/12s11' : 25.3333,
                'p9/12s12' : 26,
                'p9/12s13' : 30.6667,
                'p9/12s14' : 30.6667,
                'p9/12s15' : 33.3333,
                'p9/12s16' : 36,
                'p9/12s17' : 48.6667,
                'p9/12s18' : 48.6667,
                'p9/12s19' : 54.6667,
                'p9/12s20' : 54.6667,
                'p9/12s21' : 60.6667,
                'p9/12s22' : 60.6667,
                'p9/12s23' : 65.3333,
                'p9/12s24' : 65.3333,
                'p9/12s25' : 67.3333,
                'p9/12s26' : 72.6667,
                'p9/12s27' : 77.3333,
                'p9/12s28' : 82.6667,
                'p9/12s29' : 86,
                'p9/12s30' : 86,
                'p9/12s31' : 87,
                'p9/12s32' : 102.6667,
                'p9/12s33' : 111.3333,
                'p9/12s34' : 114,
                'p9/12s35' : 117.3333,
                'p9/12s36' : 121.3333,
                'p9/12s37' : 124.6667,
                'p9/12s38' : 127.3334,
                'p9/12s39' : 147.3334,
                'p9/12s40' : 148,
                'p10/12s10': 22,
                'p10/12s11': 26,
                'p10/12s12': 28.6667,
                'p10/12s13': 30.6667,
                'p10/12s14': 31.3333,
                'p10/12s15': 36,
                'p10/12s16': 36.6667,
                'p10/12s17': 50,
                'p10/12s18': 50,
                'p10/12s19': 58.6667,
                'p10/12s20': 58.6667,
                'p10/12s21': 60.6667,
                'p10/12s22': 63.3333,
                'p10/12s23': 65.3333,
                'p10/12s24': 69.3333,
                'p10/12s25': 72.6667,
                'p10/12s26': 78,
                'p10/12s27': 81.3333,
                'p10/12s28': 84,
                'p10/12s29': 106.6667,
                'p10/12s30': 109.3333,
                'p10/12s31': 112.6667,
                'p10/12s32': 115.3334,
                'p10/12s33': 121.3334,
                'p10/12s34': 122.6666,
                'p10/12s35': 127.3333,
                'p10/12s36': 148.6667,
                'p10/12s37': 150.6667,
                'p10/12s38': 136,
                'p10/12s39': 159.3333,
                'p10/12s40': 164,
                'p11/12s10': 24,
                'p11/12s11': 26,
                'p11/12s12': 29.3333,
                'p11/12s13': 31.3333,
                'p11/12s14': 31.3333,
                'p11/12s15': 36.6667,
                'p11/12s16': 44.6667,
                'p11/12s17': 50,
                'p11/12s18': 56.6667,
                'p11/12s19': 58.6667,
                'p11/12s20': 58.6667,
                'p11/12s21': 63.3333,
                'p11/12s22': 67.3333,
                'p11/12s23': 69.3333,
                'p11/12s24': 70.6667,
                'p11/12s25': 80.6667,
                'p11/12s26': 79.3333,
                'p11/12s27': 102.6666,
                'p11/12s28': 105.3333,
                'p11/12s29': 108.6667,
                'p11/12s30': 111.3333,
                'p11/12s31': 117.3333,
                'p11/12s32': 118.6667,
                'p11/12s33': 138,
                'p11/12s34': 141.3334,
                'p11/12s35': 150,
                'p11/12s36': 153.3334,
                'p11/12s37': 159.3334,
                'p11/12s38': 161.3333,
                'p11/12s39': 164.6667,
                'p11/12s40': 169.3333,
                'p12/12s10': 24,
                'p12/12s11': 29.3333,
                'p12/12s12': 29.3333,
                'p12/12s13': 32,
                'p12/12s14': 34.6667,
                'p12/12s15': 37.3333,
                'p12/12s16': 45.3333,
                'p12/12s17': 56.6667,
                'p12/12s18': 56.6667,
                'p12/12s19': 58.6667,
                'p12/12s20': 65.3333,
                'p12/12s21': 67.3333,
                'p12/12s22': 67.3333,
                'p12/12s23': 70.6667,
                'p12/12s24': 77.3333,
                'p12/12s25': 96,
                'p12/12s26': 98.6666,
                'p12/12s27': 102,
                'p12/12s28': 106,
                'p12/12s29': 110.6667,
                'p12/12s30': 113.3333,
                'p12/12s31': 130.3334,
                'p12/12s32': 134.6667,
                'p12/12s33': 145.3334,
                'p12/12s34': 146.6667,
                'p12/12s35': 152.3333,
                'p12/12s36': 154.6666,
                'p12/12s37': 158.6666,
                'p12/12s38': 162.6667,
                'p12/12s39': 184,
                'p12/12s40': 202.6666
                }
       
            
        
        
        bf = bfCalcDict[bfCode]
        totalBf= float(bf) * float(trussNumber)+(float(gable)*75)
        price = float(totalBf) * float(pPerBf)
        self.lPrice.setText(str(price))
        self.lTotalBf.setText(str(totalBf)+' BF') 
        
        
        
        
        print (bf)
        
        
    def bfEst(self):
        
        
        pitch = str(self.pitch.currentText())
        span = str(self.span.value())
        gable = str(self.gable.value())
        trussNumber = self.trussNumber.text()
        pPerBf = self.pPerBf.value()
        bfCode = 'p'+ pitch + 's'+ span
        
        
        if self.bcSize.currentText() == '2x4':
            bfCalcDict = {
                'p3/12s10': 15.3333,     
                'p3/12s11': 17.3333,  
                'p3/12s12': 18.6667,  
                'p3/12s13': 20,  
                'p3/12s14': 21.3333,  
                'p3/12s15': 22.6667,  
                'p3/12s16': 25.3333,  
                'p3/12s17': 30.64,  
                'p3/12s18': 33.3333, 
                'p3/12s19': 34.6667,  
                'p3/12s20': 37.3333,  
                'p3/12s21': 40,  
                'p3/12s22': 41.3333,  
                'p3/12s23': 42.6667,  
                'p3/12s24': 45.3333,  
                'p3/12s25': 48.6667,  
                'p3/12s26': 48.6667,  
                'p3/12s27': 52,  
                'p3/12s28': 54.6667,  
                'p3/12s29': 56,  
                'p3/12s30': 56,  
                'p3/12s31': 60.6667,  
                'p3/12s32': 60,  
                'p3/12s33': 68,  
                'p3/12s34': 70.6667,  
                'p3/12s35': 74.6667,  
                'p3/12s36': 86.6667,  
                'p3/12s37': 88.6667,  
                'p3/12s38': 88.6667,  
                'p3/12s39': 100.6667,  
                'p3/12s40': 99.3333,  
                'p4/12s10': 16,  
                'p4/12s11': 17.3333,  
                'p4/12s12': 18.6667,  
                'p4/12s13': 20,  
                'p4/12s14': 22,  
                'p4/12s15': 26,  
                'p4/12s16': 26,  
                'p4/12s17': 33.3333,  
                'p4/12s18': 34.6667,  
                'p4/12s19': 40,  
                'p4/12s20': 40,  
                'p4/12s21': 41.3333,  
                'p4/12s22': 41.3333,  
                'p4/12s23': 46.6667,  
                'p4/12s24': 46.6667,  
                'p4/12s25': 49.3333,  
                'p4/12s26': 52.6667,  
                'p4/12s27': 56.6667,  
                'p4/12s28': 56.6667, 
                'p4/12s29': 58,  
                'p4/12s30': 59.3333,  
                'p4/12s31':61.3333,
                'p4/12s32':69.3333,
                'p4/12s33': 74.6667,  
                'p4/12s34': 77.3333,  
                'p4/12s35': 78.6667,  
                'p4/12s36': 78.6667,  
                'p4/12s37': 81.3333,  
                'p4/12s38': 85.3333,  
                'p4/12s39': 86.6667, 
                'p4/12s40': 91.3333, 
                'p5/12s10': 16,  
                'p5/12s11': 19.3333,  
                'p5/12s12': 19.3333,  
                'p5/12s13': 22,  
                'p5/12s14': 22,  
                'p5/12s15': 26,  
                'p5/12s16': 26.6667, 
                'p5/12s17': 36,  
                'p5/12s18': 36,  
                'p5/12s19': 40,  
                'p5/12s20': 40,  
                'p5/12s21': 42.6667,  
                'p5/12s22': 45.3333,  
                'p5/12s23': 46.6667,  
                'p5/12s24': 50,  
                'p5/12s25': 52.6667,  
                'p5/12s26': 56,  
                'p5/12s27': 57.3333,  
                'p5/12s28': 57.3333, 
                'p5/12s29': 60,  
                'p5/12s30': 62.6667,  
                'p5/12s31':64,
                'p5/12s32': 74.6667,  
                'p5/12s33': 77.3333,  
                'p5/12s34': 81.3333,  
                'p5/12s35':82.6667,
                'p5/12s36': 85.3333,  
                'p5/12s37':89.3333,
                'p5/12s38':90.6667,
                'p5/12s39': 93.3333, 
                'p5/12s40': 97.3333,  
                'p6/12s10': 16.6667,  
                'p6/12s11': 19.3333,  
                'p6/12s12': 19.3333,  
                'p6/12s13': 22,  
                'p6/12s14': 22.6667,  
                'p6/12s15': 26.6667, 
                'p6/12s16': 26.6667, 
                'p6/12s17': 36,  
                'p6/12s18': 40,  
                'p6/12s19': 41.3333,  
                'p6/12s20': 41.3333,  
                'p6/12s21': 44,  
                'p6/12s22': 48,  
                'p6/12s23': 49.3333,  
                'p6/12s24': 52,  
                'p6/12s25': 57.3333,  
                'p6/12s26': 58,  
                'p6/12s27': 60.6667,  
                'p6/12s28': 60.6667,  
                'p6/12s29': 63.3333,  
                'p6/12s30': 64,  
                'p6/12s31': 70,  
                'p6/12s32': 82.6667,  
                'p6/12s33': 84,  
                'p6/12s34':85.3333,
                'p6/12s35': 89.3333,  
                'p6/12s36': 92,  
                'p6/12s37': 93.3333,  
                'p6/12s38': 93.3333,  
                'p6/12s39': 100,  
                'p6/12s40': 104,  
                'p7/12s10': 16.6667,  
                'p7/12s11': 19.3333,  
                'p7/12s12': 21.3333,  
                'p7/12s13': 22.6667,  
                'p7/12s14': 25.3333,  
                'p7/12s15': 27.3333,  
                'p7/12s16': 36,  
                'p7/12s17': 40, 
                'p7/12s18': 40,  
                'p7/12s19': 44,  
                'p7/12s20': 44,  
                'p7/12s21': 48,  
                'p7/12s22': 49.3333,  
                'p7/12s23': 50.6667,  
                'p7/12s24': 56.6667,  
                'p7/12s25': 58.6667,  
                'p7/12s26':60,
                'p7/12s27':61.3333,
                'p7/12s28': 64,  
                'p7/12s29': 66.6667,  
                'p7/12s30': 72,  
                'p7/12s31': 86.6667,  
                'p7/12s32': 86.6667,  
                'p7/12s33': 88,  
                'p7/12s34': 90.6667,  
                'p7/12s35': 98.6667,  
                'p7/12s36': 98.6667,  
                'p7/12s37': 100,  
                'p7/12s38': 102.6667,  
                'p7/12s39': 106.6667,
                'p7/12s40': 106.6667,    
                'p8/12s10': 18.6667,  
                'p8/12s11': 20,  
                'p8/12s12': 21.3333,  
                'p8/12s13': 26,  
                'p8/12s14': 26, 
                'p8/12s15': 27.3333,  
                'p8/12s16': 36,  
                'p8/12s17': 42.6667,  
                'p8/12s18': 42.6667,  
                'p8/12s19': 44,  
                'p8/12s20': 48,  
                'p8/12s21': 49.3333,  
                'p8/12s22': 49.3333,  
                'p8/12s23': 57.3333,  
                'p8/12s24': 57.3333, 
                'p8/12s25': 62.6667,  
                'p8/12s26':62.6667,
                'p8/12s27': 65.3333,  
                'p8/12s28': 68,  
                'p8/12s29': 70.6667,  
                'p8/12s30': 73.3333,  
                'p8/12s31':90.6667,
                'p8/12s32':90.6667,
                'p8/12s33': 97.3333,  
                'p8/12s34': 97.3333,  
                'p8/12s35': 102.6667,  
                'p8/12s36': 117.3333,  
                'p8/12s37': 121.3333,  
                'p8/12s38': 124,  
                'p8/12s39': 126.6667,  
                'p8/12s40': 130.6667,  
                'p9/12s10': 18.6667,  
                'p9/12s11': 21.3333,  
                'p9/12s12': 22,  
                'p9/12s13': 26, 
                'p9/12s14': 26, 
                'p9/12s15': 28,  
                'p9/12s16': 41.3333,  
                'p9/12s17': 42.6667, 
                'p9/12s18': 44,  
                'p9/12s19': 48,  
                'p9/12s20': 48,  
                'p9/12s21': 53.3333,  
                'p9/12s22': 53.3333,  
                'p9/12s23': 57.3333,  
                'p9/12s24': 61.3333, 
                'p9/12s25': 62.6667,  
                'p9/12s26': 64,  
                'p9/12s27':72,
                'p9/12s28': 72,  
                'p9/12s29': 76, 
                'p9/12s30': 76,  
                'p9/12s31': 92,  
                'p9/12s32': 102.6667,  
                'p9/12s33': 116,  
                'p9/12s34': 118.6667,  
                'p9/12s35': 122.6667,  
                'p9/12s36': 125.3334,  
                'p9/12s37': 129.3334,  
                'p9/12s38': 130.6666,  
                'p9/12s39': 150,  
                'p9/12s40': 154.6666,  
                'p10/12s10':18.6667,  
                'p10/12s11':22,  
                'p10/12s12':24.6667,  
                'p10/12s13':26.6667, 
                'p10/12s14':26.6667, 
                'p10/12s15':30.6667,  
                'p10/12s16':41.3333,  
                'p10/12s17':44, 
                'p10/12s18':44,  
                'p10/12s19':52,  
                'p10/12s20':52,  
                'p10/12s21':53.3333,  
                'p10/12s22':56, 
                'p10/12s23':61.3333,  
                'p10/12s24':61.3333,  
                'p10/12s25':66.6667,  
                'p10/12s26':72,  
                'p10/12s27':72,  
                'p10/12s28':74.6667,  
                'p10/12s29':141.3333,  
                'p10/12s30':98,   
                'p10/12s31':100.6667,   
                'p10/12s32':103.3334,   
                'p10/12s33':107.3333,   
                'p10/12s34':110,  
                'p10/12s35':114,   
                'p10/12s36':130.6667,   
                'p10/12s37':140,   
                'p10/12s38':142.6667,   
                'p10/12s39':145.3333,   
                'p10/12s40':148.6667,  
                'p11/12s10':20.6667,  
                'p11/12s11':22,  
                'p11/12s12':25.3333,  
                'p11/12s13':26.6667, 
                'p11/12s14':27.3333,  
                'p11/12s15':31.3333,  
                'p11/12s16':42.6667,  
                'p11/12s17':48, 
                'p11/12s18':50.6667,  
                'p11/12s19':52,  
                'p11/12s20':52,  
                'p11/12s21':57.3333,  
                'p11/12s22':60,  
                'p11/12s23':61.3333,  
                'p11/12s24':65.3333,  
                'p11/12s25':72, 
                'p11/12s26':70.6667, 
                'p11/12s27':92,   
                'p11/12s28':94.6667,   
                'p11/12s29':98.6667,   
                'p11/12s30':102.6667,   
                'p11/12s31':106.6667,   
                'p11/12s32':109.3334,   
                'p11/12s33':126.6666,   
                'p11/12s34':132.6666,   
                'p11/12s35':139.3334,   
                'p11/12s36':140,   
                'p11/12s37':144,   
                'p11/12s38':150,  
                'p11/12s39':152.6667,   
                'p11/12s40':154.6667,   
                'p12/12s10':20.6667,  
                'p12/12s11':25.3333,  
                'p12/12s12':25.3333,  
                'p12/12s13':27.3333,  
                'p12/12s14':30, 
                'p12/12s15':32,  
                'p12/12s16':46.6667,  
                'p12/12s17':50.6667,  
                'p12/12s18':50.6667,  
                'p12/12s19':52,  
                'p12/12s20':58.6667, 
                'p12/12s21':60,  
                'p12/12s22':60,  
                'p12/12s23':62.6667,  
                'p12/12s24':72,  
                'p12/12s25':86,  
                'p12/12s26':88.6667, 
                'p12/12s27':92.6667,   
                'p12/12s28':96.6667,   
                'p12/12s29':99.3334,   
                'p12/12s30':103.3334,  
                'p12/12s31':121.3334,  
                'p12/12s32':126.6666,   
                'p12/12s33':134,   
                'p12/12s34':136.6667,   
                'p12/12s35':142,   
                'p12/12s36':144,   
                'p12/12s37':147.3333,   
                'p12/12s38':150,   
                'p12/12s39':172.6667,   
                'p12/12s40':188.6667,   
                
                
            }
            
        if self.bcSize.currentText() == '2x6':
        
        
            bfCalcDict = {

                'p3/12s10' : 18,
                'p3/12s11' : 20,
                'p3/12s12' : 22.6667,
                'p3/12s13' : 24.6667,
                'p3/12s14' : 26,
                'p3/12s15' : 28,
                'p3/12s16' : 30.6667,
                'p3/12s17' : 38,
                'p3/12s18' : 39.3333,
                'p3/12s19' : 41.6667,
                'p3/12s20' : 44,
                'p3/12s21' : 46,
                'p3/12s22' : 47.3333,
                'p3/12s23' : 50.6667,
                'p3/12s24' : 53.3333,
                'p3/12s25' : 55.3333,
                'p3/12s26' : 55.3333,
                'p3/12s27' : 57.3333,
                'p3/12s28' : 64,
                'p3/12s29' : 66,
                'p3/12s30' : 66,
                'p3/12s31' : 69.6667,
                'p3/12s32' : 72,
                'p3/12s33' : 74,
                'p3/12s34' : 79.3333,
                'p3/12s35' : 84,
                'p3/12s36' : 86.6667,
                'p3/12s37' : 88.6667,
                'p3/12s38' : 88.6667,
                'p3/12s39' : 110,
                'p3/12s40' : 122,
                'p4/12s10' : 18,
                'p4/12s11' : 21.3333,
                'p4/12s12' : 22.6667,
                'p4/12s13' : 24.6667,
                'p4/12s14' : 26,
                'p4/12s15' : 31.3333,
                'p4/12s16' : 31.3333,
                'p4/12s17' : 39.3333,
                'p4/12s18' : 39.3333,
                'p4/12s19' : 45.3333,
                'p4/12s20' : 46.6667,
                'p4/12s21' : 48.6667,
                'p4/12s22' : 48.6667,
                'p4/12s23' : 53.3333,
                'p4/12s24' : 54.6667,
                'p4/12s25' : 56.6667,
                'p4/12s26' : 56.6667,
                'p4/12s27' : 66,
                'p4/12s28' : 66,
                'p4/12s29' : 68,
                'p4/12s30' : 68,
                'p4/12s31' : 73,
                'p4/12s32' : 72.6667,
                'p4/12s33' : 82,
                'p4/12s34' : 86,
                'p4/12s35' : 90.6667,
                'p4/12s36' : 90.6667,
                'p4/12s37' : 92.6667,
                'p4/12s38' : 96.6667,
                'p4/12s39' : 100,
                'p4/12s40' : 100,
                'p5/12s10' : 18,
                'p5/12s11' : 22.6667,
                'p5/12s12' : 23.3333,
                'p5/12s13' : 26.6667,
                'p5/12s14' : 26.6667,
                'p5/12s15' : 31.3333,
                'p5/12s16' : 31.3333,
                'p5/12s17' : 40.6667,
                'p5/12s18' : 42,
                'p5/12s19' : 46.6667,
                'p5/12s20' : 46.6667,
                'p5/12s21' : 50,
                'p5/12s22' : 52.6667,
                'p5/12s23' : 54.6667,
                'p5/12s24' : 54.6667,
                'p5/12s25' : 59.3333,
                'p5/12s26' : 62,
                'p5/12s27' : 66.6667,
                'p5/12s28' : 66.6667,
                'p5/12s29' : 70,
                'p5/12s30' : 71.3333,
                'p5/12s31' : 74.3333,
                'p5/12s32' : 78,
                'p5/12s33' : 85.3333,
                'p5/12s34' : 88,
                'p5/12s35' : 94.6667,
                'p5/12s36' : 94.6667,
                'p5/12s37' : 102,
                'p5/12s38' : 102,
                'p5/12s39' : 105.3333,
                'p5/12s40' : 105.3333,
                'p6/12s10' : 20,
                'p6/12s11' : 23.3333,
                'p6/12s12' : 23.3333,
                'p6/12s13' : 26.6667,
                'p6/12s14' : 27.3333,
                'p6/12s15' : 32,
                'p6/12s16' : 32,
                'p6/12s17' : 42,
                'p6/12s18' : 44.6667,
                'p6/12s19' : 48,
                'p6/12s20' : 48,
                'p6/12s21' : 50,
                'p6/12s22' : 55.3333,
                'p6/12s23' : 57.3333,
                'p6/12s24' : 57.3333,
                'p6/12s25' : 62,
                'p6/12s26' : 66.6667,
                'p6/12s27' : 68.6667,
                'p6/12s28' : 70,
                'p6/12s29' : 73.3333,
                'p6/12s30' : 74,
                'p6/12s31' : 77.6667,
                'p6/12s32' : 88,
                'p6/12s33' : 90,
                'p6/12s34' : 95.3333,
                'p6/12s35' : 98.6667,
                'p6/12s36' : 104,
                'p6/12s37' : 106,
                'p6/12s38' : 106,
                'p6/12s39' : 110.6667,
                'p6/12s40' : 116,
                'p7/12s10' : 20,
                'p7/12s11' : 23.3333,
                'p7/12s12' : 25.3333,
                'p7/12s13' : 27.3333,
                'p7/12s14' : 30,
                'p7/12s15' : 32,
                'p7/12s16' : 32.6667,
                'p7/12s17' : 46,
                'p7/12s18' : 46,
                'p7/12s19' : 48,
                'p7/12s20' : 50.6667,
                'p7/12s21' : 55.3333,
                'p7/12s22' : 55.3333,
                'p7/12s23' : 58.6667,
                'p7/12s24' : 61.3333,
                'p7/12s25' : 63.3333,
                'p7/12s26' : 67.3333,
                'p7/12s27' : 70.6667,
                'p7/12s28' : 72,
                'p7/12s29' : 76.6667,
                'p7/12s30' : 82,
                'p7/12s31' : 84.3333,
                'p7/12s32' : 90.6667,
                'p7/12s33' : 94,
                'p7/12s34' : 102,
                'p7/12s35' : 106.6667,
                'p7/12s36' : 110.6667,
                'p7/12s37' : 112.6667,
                'p7/12s38' : 115.3333,
                'p7/12s39' : 117.3333,
                'p7/12s40' : 117.3333,
                'p8/12s10' : 21.3333,
                'p8/12s11' : 24,
                'p8/12s12' : 25.3333,
                'p8/12s13' : 30,
                'p8/12s14' : 30.6667,
                'p8/12s15' : 32.6667,
                'p8/12s16' : 32.6667,
                'p8/12s17' : 46,
                'p8/12s18' : 48.6667,
                'p8/12s19' : 50.6667,
                'p8/12s20' : 53.3333,
                'p8/12s21' : 56.6667,
                'p8/12s22' : 56.6667,
                'p8/12s23' : 65.3333,
                'p8/12s24' : 65.3333,
                'p8/12s25' : 67.3333,
                'p8/12s26' : 71.3333,
                'p8/12s27' : 74.6667,
                'p8/12s28' : 77.3333,
                'p8/12s29' : 82,
                'p8/12s30' : 83.3333,
                'p8/12s31' : 85.6667,
                'p8/12s32' : 97.3333,
                'p8/12s33' : 102,
                'p8/12s34' : 108.6667,
                'p8/12s35' : 112,
                'p8/12s36' : 129.3333,
                'p8/12s37' : 135.3333,
                'p8/12s38' : 138,
                'p8/12s39' : 141.3333,
                'p8/12s40' : 145.3333,
                'p9/12s10' : 22,
                'p9/12s11' : 25.3333,
                'p9/12s12' : 26,
                'p9/12s13' : 30.6667,
                'p9/12s14' : 30.6667,
                'p9/12s15' : 33.3333,
                'p9/12s16' : 36,
                'p9/12s17' : 48.6667,
                'p9/12s18' : 48.6667,
                'p9/12s19' : 54.6667,
                'p9/12s20' : 54.6667,
                'p9/12s21' : 60.6667,
                'p9/12s22' : 60.6667,
                'p9/12s23' : 65.3333,
                'p9/12s24' : 65.3333,
                'p9/12s25' : 67.3333,
                'p9/12s26' : 72.6667,
                'p9/12s27' : 77.3333,
                'p9/12s28' : 82.6667,
                'p9/12s29' : 86,
                'p9/12s30' : 86,
                'p9/12s31' : 87,
                'p9/12s32' : 102.6667,
                'p9/12s33' : 111.3333,
                'p9/12s34' : 114,
                'p9/12s35' : 117.3333,
                'p9/12s36' : 121.3333,
                'p9/12s37' : 124.6667,
                'p9/12s38' : 127.3334,
                'p9/12s39' : 147.3334,
                'p9/12s40' : 148,
                'p10/12s10': 22,
                'p10/12s11': 26,
                'p10/12s12': 28.6667,
                'p10/12s13': 30.6667,
                'p10/12s14': 31.3333,
                'p10/12s15': 36,
                'p10/12s16': 36.6667,
                'p10/12s17': 50,
                'p10/12s18': 50,
                'p10/12s19': 58.6667,
                'p10/12s20': 58.6667,
                'p10/12s21': 60.6667,
                'p10/12s22': 63.3333,
                'p10/12s23': 65.3333,
                'p10/12s24': 69.3333,
                'p10/12s25': 72.6667,
                'p10/12s26': 78,
                'p10/12s27': 81.3333,
                'p10/12s28': 84,
                'p10/12s29': 106.6667,
                'p10/12s30': 109.3333,
                'p10/12s31': 112.6667,
                'p10/12s32': 115.3334,
                'p10/12s33': 121.3334,
                'p10/12s34': 122.6666,
                'p10/12s35': 127.3333,
                'p10/12s36': 148.6667,
                'p10/12s37': 150.6667,
                'p10/12s38': 136,
                'p10/12s39': 159.3333,
                'p10/12s40': 164,
                'p11/12s10': 24,
                'p11/12s11': 26,
                'p11/12s12': 29.3333,
                'p11/12s13': 31.3333,
                'p11/12s14': 31.3333,
                'p11/12s15': 36.6667,
                'p11/12s16': 44.6667,
                'p11/12s17': 50,
                'p11/12s18': 56.6667,
                'p11/12s19': 58.6667,
                'p11/12s20': 58.6667,
                'p11/12s21': 63.3333,
                'p11/12s22': 67.3333,
                'p11/12s23': 69.3333,
                'p11/12s24': 70.6667,
                'p11/12s25': 80.6667,
                'p11/12s26': 79.3333,
                'p11/12s27': 102.6666,
                'p11/12s28': 105.3333,
                'p11/12s29': 108.6667,
                'p11/12s30': 111.3333,
                'p11/12s31': 117.3333,
                'p11/12s32': 118.6667,
                'p11/12s33': 138,
                'p11/12s34': 141.3334,
                'p11/12s35': 150,
                'p11/12s36': 153.3334,
                'p11/12s37': 159.3334,
                'p11/12s38': 161.3333,
                'p11/12s39': 164.6667,
                'p11/12s40': 169.3333,
                'p12/12s10': 24,
                'p12/12s11': 29.3333,
                'p12/12s12': 29.3333,
                'p12/12s13': 32,
                'p12/12s14': 34.6667,
                'p12/12s15': 37.3333,
                'p12/12s16': 45.3333,
                'p12/12s17': 56.6667,
                'p12/12s18': 56.6667,
                'p12/12s19': 58.6667,
                'p12/12s20': 65.3333,
                'p12/12s21': 67.3333,
                'p12/12s22': 67.3333,
                'p12/12s23': 70.6667,
                'p12/12s24': 77.3333,
                'p12/12s25': 96,
                'p12/12s26': 98.6666,
                'p12/12s27': 102,
                'p12/12s28': 106,
                'p12/12s29': 110.6667,
                'p12/12s30': 113.3333,
                'p12/12s31': 130.3334,
                'p12/12s32': 134.6667,
                'p12/12s33': 145.3334,
                'p12/12s34': 146.6667,
                'p12/12s35': 152.3333,
                'p12/12s36': 154.6666,
                'p12/12s37': 158.6666,
                'p12/12s38': 162.6667,
                'p12/12s39': 184,
                'p12/12s40': 202.6666
                }
       
            
        
        
        bf = bfCalcDict[bfCode]
        totalBf= float(bf) * float(trussNumber)+(float(gable)*75)
        price = float(totalBf) * float(pPerBf)
        self.lPrice.setText(str(price))
        self.lTotalBf.setText(str(totalBf)+' BF') 
        
        
        
        
        print (bf)
        
    def tax(self):
        quotedPrice = self.inputQuotedPrice.text()
        zip = self.inputZipCode.text()
        zipdict = {
            '30002':0.08,
            '30003':0.06,
            '30004':0.0775,
            '30005':0.0775,
            '30006':0.06,
            '30007':0.06,
            '30008':0.06,
            '30009':0.0775,
            '30010':0.06,
            '30011':0.07,
            '30012':0.07,
            '30013':0.07,
            '30014':0.07,
            '30015':0.07,
            '30016':0.07,
            '30017':0.06,
            '30018':0.07,
            '30019':0.06,
            '30021':0.08,
            '30022':0.0775,
            '30023':0.0775,
            '30024':0.06,
            '30025':0.07,
            '30026':0.06,
            '30028':0.07,
            '30029':0.06,
            '30030':0.08,
            '30031':0.08,
            '30032':0.08,
            '30033':0.08,
            '30034':0.08,
            '30035':0.08,
            '30036':0.08,
            '30037':0.08,
            '30038':0.08,
            '30039':0.06,
            '30040':0.07,
            '30041':0.07,
            '30042':0.06,
            '30043':0.06,
            '30044':0.06,
            '30045':0.06,
            '30046':0.06,
            '30047':0.06,
            '30048':0.06,
            '30049':0.06,
            '30052':0.06,
            '30054':0.07,
            '30055':0.07,
            '30056':0.07,
            '30058':0.08,
            '30060':0.06,
            '30061':0.06,
            '30062':0.06,
            '30063':0.06,
            '30064':0.06,
            '30065':0.06,
            '30066':0.06,
            '30067':0.06,
            '30068':0.06,
            '30069':0.06,
            '30070':0.07,
            '30071':0.06,
            '30072':0.08,
            '30074':0.08,
            '30075':0.0775,
            '30076':0.0775,
            '30077':0.0775,
            '30078':0.06,
            '30079':0.08,
            '30080':0.06,
            '30081':0.06,
            '30082':0.06,
            '30083':0.08,
            '30084':0.08,
            '30085':0.08,
            '30086':0.08,
            '30087':0.08,
            '30088':0.08,
            '30090':0.06,
            '30091':0.06,
            '30092':0.06,
            '30093':0.06,
            '30094':0.07,
            '30095':0.06,
            '30096':0.06,
            '30097':0.06,
            '30098':0.0775,
            '30099':0.06,
            '30101':0.06,
            '30102':0.06,
            '30103':0.07,
            '30104':0.07,
            '30105':0.07,
            '30106':0.06,
            '30107':0.06,
            '30108':0.07,
            '30109':0.07,
            '30110':0.08,
            '30111':0.06,
            '30112':0.07,
            '30113':0.08,
            '30114':0.06,
            '30115':0.06,
            '30116':0.07,
            '30117':0.07,
            '30118':0.07,
            '30119':0.07,
            '30120':0.07,
            '30121':0.07,
            '30122':0.07,
            '30123':0.07,
            '30124':0.07,
            '30125':0.07,
            '30126':0.06,
            '30127':0.06,
            '30129':0.07,
            '30132':0.07,
            '30133':0.07,
            '30134':0.07,
            '30135':0.07,
            '30137':0.07,
            '30138':0.07,
            '30139':0.07,
            '30140':0.08,
            '30141':0.07,
            '30142':0.06,
            '30143':0.07,
            '30144':0.06,
            '30145':0.07,
            '30146':0.06,
            '30147':0.07,
            '30148':0.07,
            '30149':0.07,
            '30150':0.07,
            '30151':0.06,
            '30152':0.06,
            '30153':0.07,
            '30154':0.07,
            '30156':0.06,
            '30157':0.07,
            '30160':0.06,
            '30161':0.07,
            '30162':0.07,
            '30164':0.07,
            '30165':0.07,
            '30168':0.06,
            '30169':0.06,
            '30170':0.07,
            '30171':0.07,
            '30172':0.07,
            '30173':0.07,
            '30175':0.07,
            '30176':0.08,
            '30177':0.07,
            '30178':0.07,
            '30179':0.07,
            '30180':0.07,
            '30182':0.07,
            '30183':0.06,
            '30184':0.07,
            '30185':0.07,
            '30187':0.07,
            '30188':0.06,
            '30189':0.06,
            '30204':0.07,
            '30205':0.07,
            '30206':0.07,
            '30212':0.07,
            '30213':0.0775,
            '30214':0.07,
            '30215':0.07,
            '30216':0.07,
            '30217':0.07,
            '30218':0.07,
            '30219':0.07,
            '30220':0.07,
            '30222':0.07,
            '30223':0.07,
            '30224':0.07,
            '30228':0.07,
            '30229':0.07,
            '30230':0.07,
            '30233':0.07,
            '30234':0.07,
            '30236':0.08,
            '30237':0.08,
            '30238':0.08,
            '30240':0.07,
            '30241':0.07,
            '30248':0.07,
            '30250':0.08,
            '30251':0.07,
            '30252':0.07,
            '30253':0.07,
            '30256':0.07,
            '30257':0.07,
            '30258':0.07,
            '30259':0.07,
            '30260':0.08,
            '30261':0.07,
            '30263':0.07,
            '30264':0.07,
            '30265':0.07,
            '30266':0.07,
            '30268':0.0775,
            '30269':0.07,
            '30270':0.07,
            '30271':0.07,
            '30272':0.0775,
            '30273':0.08,
            '30274':0.08,
            '30275':0.07,
            '30276':0.07,
            '30277':0.07,
            '30281':0.07,
            '30284':0.07,
            '30285':0.07,
            '30286':0.07,
            '30287':0.08,
            '30288':0.08,
            '30289':0.07,
            '30290':0.07,
            '30291':0.0775,
            '30292':0.07,
            '30293':0.07,
            '30294':0.08,
            '30295':0.07,
            '30296':0.08,
            '30297':0.08,
            '30298':0.08,
            '30301':0.0775,
            '30302':0.0775,
            '30303':0.089,
            '30304':0.089,
            '30305':0.089,
            '30306':0.089,
            '30307':0.089,
            '30308':0.089,
            '30309':0.089,
            '30310':0.089,
            '30311':0.089,
            '30312':0.089,
            '30313':0.089,
            '30314':0.089,
            '30315':0.089,
            '30316':0.08,
            '30317':0.089,
            '30318':0.089,
            '30319':0.08,
            '30320':0.0775,
            '30321':0.089,
            '30322':0.08,
            '30324':0.089,
            '30325':0.0775,
            '30326':0.089,
            '30327':0.089,
            '30328':0.0775,
            '30329':0.08,
            '30331':0.089,
            '30332':0.0775,
            '30333':0.08,
            '30334':0.0775,
            '30336':0.0775,
            '30337':0.0775,
            '30338':0.08,
            '30339':0.06,
            '30340':0.08,
            '30341':0.08,
            '30342':0.089,
            '30343':0.0775,
            '30344':0.0775,
            '30345':0.08,
            '30346':0.08,
            '30348':0.0775,
            '30349':0.0775,
            '30350':0.0775,
            '30353':0.0775,
            '30354':0.089,
            '30355':0.089,
            '30356':0.08,
            '30357':0.0775,
            '30358':0.0775,
            '30359':0.08,
            '30360':0.08,
            '30361':0.089,
            '30362':0.08,
            '30363':0.089,
            '30364':0.0775,
            '30366':0.08,
            '30368':0.0775,
            '30369':0.089,
            '30370':0.089,
            '30371':0.089,
            '30374':0.089,
            '30375':0.089,
            '30377':0.0775,
            '30378':0.0775,
            '30380':0.0775,
            '30384':0.0775,
            '30385':0.0775,
            '30388':0.0775,
            '30392':0.0775,
            '30394':0.0775,
            '30396':0.0775,
            '30398':0.0775,
            '30401':0.08,
            '30410':0.08,
            '30411':0.08,
            '30412':0.08,
            '30413':0.08,
            '30414':0.08,
            '30415':0.07,
            '30417':0.08,
            '30420':0.08,
            '30421':0.08,
            '30423':0.08,
            '30424':0.07,
            '30425':0.07,
            '30426':0.07,
            '30427':0.08,
            '30428':0.08,
            '30429':0.08,
            '30434':0.08,
            '30436':0.08,
            '30438':0.08,
            '30439':0.08,
            '30441':0.07,
            '30442':0.08,
            '30445':0.08,
            '30446':0.07,
            '30447':0.08,
            '30448':0.08,
            '30449':0.07,
            '30450':0.07,
            '30451':0.08,
            '30452':0.07,
            '30453':0.08,
            '30454':0.08,
            '30455':0.07,
            '30456':0.07,
            '30457':0.08,
            '30458':0.07,
            '30459':0.07,
            '30460':0.07,
            '30461':0.07,
            '30464':0.08,
            '30467':0.07,
            '30470':0.08,
            '30471':0.08,
            '30473':0.08,
            '30474':0.08,
            '30475':0.08,
            '30477':0.08,
            '30499':0.08,
            '30501':0.07,
            '30502':0.07,
            '30503':0.07,
            '30504':0.07,
            '30506':0.07,
            '30507':0.07,
            '30510':0.07,
            '30511':0.07,
            '30512':0.07,
            '30513':0.07,
            '30514':0.07,
            '30515':0.06,
            '30516':0.07,
            '30517':0.07,
            '30518':0.06,
            '30519':0.06,
            '30520':0.07,
            '30521':0.07,
            '30522':0.07,
            '30523':0.07,
            '30525':0.07,
            '30527':0.07,
            '30528':0.07,
            '30529':0.07,
            '30530':0.07,
            '30531':0.07,
            '30533':0.07,
            '30534':0.07,
            '30535':0.07,
            '30536':0.07,
            '30537':0.07,
            '30538':0.07,
            '30539':0.07,
            '30540':0.07,
            '30541':0.07,
            '30542':0.07,
            '30543':0.07,
            '30545':0.07,
            '30546':0.07,
            '30547':0.07,
            '30548':0.07,
            '30549':0.07,
            '30552':0.07,
            '30553':0.07,
            '30554':0.07,
            '30555':0.07,
            '30557':0.07,
            '30558':0.07,
            '30559':0.07,
            '30560':0.07,
            '30562':0.07,
            '30563':0.07,
            '30564':0.07,
            '30565':0.07,
            '30566':0.07,
            '30567':0.07,
            '30568':0.07,
            '30571':0.07,
            '30572':0.07,
            '30573':0.07,
            '30575':0.07,
            '30576':0.07,
            '30577':0.07,
            '30580':0.07,
            '30581':0.07,
            '30582':0.07,
            '30597':0.07,
            '30598':0.07,
            '30599':0.07,
            '30601':0.08,
            '30602':0.08,
            '30603':0.08,
            '30604':0.08,
            '30605':0.08,
            '30606':0.08,
            '30607':0.07,
            '30608':0.08,
            '30609':0.08,
            '30612':0.08,
            '30619':0.07,
            '30620':0.07,
            '30621':0.07,
            '30622':0.07,
            '30623':0.07,
            '30624':0.07,
            '30625':0.07,
            '30627':0.07,
            '30628':0.07,
            '30629':0.07,
            '30630':0.07,
            '30631':0.08,
            '30633':0.07,
            '30634':0.07,
            '30635':0.07,
            '30638':0.07,
            '30639':0.07,
            '30641':0.07,
            '30642':0.07,
            '30643':0.07,
            '30645':0.07,
            '30646':0.07,
            '30647':0.07,
            '30648':0.07,
            '30650':0.07,
            '30655':0.07,
            '30656':0.07,
            '30660':0.08,
            '30662':0.07,
            '30663':0.07,
            '30664':0.08,
            '30665':0.07,
            '30666':0.07,
            '30667':0.07,
            '30668':0.08,
            '30669':0.07,
            '30671':0.07,
            '30673':0.08,
            '30677':0.07,
            '30678':0.07,
            '30680':0.07,
            '30683':0.08,
            '30701':0.07,
            '30703':0.07,
            '30705':0.07,
            '30707':0.08,
            '30708':0.07,
            '30710':0.07,
            '30711':0.07,
            '30719':0.07,
            '30720':0.07,
            '30721':0.07,
            '30722':0.07,
            '30724':0.07,
            '30725':0.08,
            '30726':0.07,
            '30728':0.08,
            '30730':0.07,
            '30731':0.07,
            '30732':0.07,
            '30733':0.07,
            '30734':0.07,
            '30735':0.07,
            '30736':0.07,
            '30738':0.07,
            '30739':0.08,
            '30740':0.07,
            '30741':0.08,
            '30742':0.07,
            '30746':0.07,
            '30747':0.07,
            '30750':0.08,
            '30751':0.07,
            '30752':0.07,
            '30753':0.07,
            '30755':0.07,
            '30756':0.07,
            '30757':0.07,
            '30802':0.08,
            '30803':0.08,
            '30805':0.07,
            '30806':0.08,
            '30807':0.08,
            '30808':0.08,
            '30809':0.08,
            '30810':0.08,
            '30811':0.07,
            '30812':0.08,
            '30813':0.08,
            '30814':0.08,
            '30815':0.08,
            '30816':0.07,
            '30817':0.08,
            '30818':0.08,
            '30819':0.08,
            '30820':0.08,
            '30821':0.08,
            '30822':0.08,
            '30823':0.08,
            '30824':0.08,
            '30828':0.08,
            '30830':0.07,
            '30833':0.08,
            '30901':0.08,
            '30903':0.08,
            '30904':0.08,
            '30905':0.08,
            '30906':0.08,
            '30907':0.08,
            '30909':0.08,
            '30912':0.08,
            '30914':0.08,
            '30916':0.08,
            '30917':0.08,
            '30919':0.08,
            '30999':0.08,
            '31001':0.08,
            '31002':0.08,
            '31003':0.07,
            '31004':0.07,
            '31005':0.07,
            '31006':0.08,
            '31007':0.08,
            '31008':0.07,
            '31009':0.08,
            '31010':0.08,
            '31011':0.08,
            '31012':0.08,
            '31013':0.07,
            '31014':0.08,
            '31015':0.08,
            '31016':0.07,
            '31017':0.07,
            '31018':0.08,
            '31019':0.08,
            '31020':0.07,
            '31021':0.08,
            '31022':0.08,
            '31023':0.08,
            '31024':0.08,
            '31025':0.07,
            '31026':0.08,
            '31027':0.08,
            '31028':0.07,
            '31029':0.07,
            '31030':0.07,
            '31031':0.07,
            '31032':0.07,
            '31033':0.07,
            '31034':0.07,
            '31035':0.08,
            '31036':0.07,
            '31037':0.08,
            '31038':0.07,
            '31039':0.08,
            '31040':0.08,
            '31041':0.08,
            '31042':0.07,
            '31044':0.07,
            '31045':0.08,
            '31046':0.07,
            '31047':0.07,
            '31049':0.08,
            '31050':0.07,
            '31051':0.08,
            '31052':0.07,
            '31054':0.07,
            '31055':0.08,
            '31057':0.08,
            '31058':0.08,
            '31059':0.07,
            '31060':0.08,
            '31061':0.07,
            '31062':0.07,
            '31063':0.08,
            '31064':0.07,
            '31065':0.08,
            '31066':0.07,
            '31067':0.08,
            '31068':0.08,
            '31069':0.07,
            '31070':0.08,
            '31071':0.08,
            '31072':0.08,
            '31075':0.08,
            '31076':0.08,
            '31077':0.08,
            '31078':0.07,
            '31079':0.08,
            '31081':0.08,
            '31082':0.08,
            '31083':0.08,
            '31084':0.08,
            '31085':0.07,
            '31086':0.07,
            '31087':0.08,
            '31088':0.07,
            '31089':0.08,
            '31090':0.07,
            '31091':0.08,
            '31092':0.08,
            '31093':0.07,
            '31094':0.08,
            '31095':0.07,
            '31096':0.08,
            '31097':0.07,
            '31098':0.07,
            '31099':0.07,
            '31106':0.0775,
            '31107':0.0775,
            '31119':0.08,
            '31126':0.0775,
            '31131':0.0775,
            '31136':0.0775,
            '31139':0.0775,
            '31141':0.08,
            '31144':0.06,
            '31145':0.08,
            '31146':0.08,
            '31150':0.0775,
            '31156':0.0775,
            '31169':0.07,
            '31192':0.0775,
            '31193':0.0775,
            '31195':0.0775,
            '31196':0.0775,
            '31201':0.07,
            '31202':0.07,
            '31203':0.07,
            '31204':0.07,
            '31205':0.07,
            '31206':0.07,
            '31207':0.07,
            '31208':0.07,
            '31209':0.07,
            '31210':0.07,
            '31211':0.07,
            '31213':0.07,
            '31216':0.07,
            '31217':0.07,
            '31220':0.07,
            '31221':0.07,
            '31294':0.07,
            '31295':0.07,
            '31296':0.07,
            '31301':0.07,
            '31302':0.07,
            '31303':0.07,
            '31304':0.07,
            '31305':0.07,
            '31307':0.07,
            '31308':0.07,
            '31309':0.07,
            '31310':0.07,
            '31312':0.07,
            '31313':0.07,
            '31314':0.07,
            '31315':0.07,
            '31316':0.07,
            '31318':0.07,
            '31319':0.07,
            '31320':0.07,
            '31321':0.07,
            '31322':0.07,
            '31323':0.07,
            '31324':0.07,
            '31326':0.07,
            '31327':0.07,
            '31328':0.07,
            '31329':0.07,
            '31331':0.07,
            '31333':0.07,
            '31401':0.07,
            '31402':0.07,
            '31403':0.07,
            '31404':0.07,
            '31405':0.07,
            '31406':0.07,
            '31407':0.07,
            '31408':0.07,
            '31409':0.07,
            '31410':0.07,
            '31411':0.07,
            '31412':0.07,
            '31414':0.07,
            '31415':0.07,
            '31416':0.07,
            '31418':0.07,
            '31419':0.07,
            '31420':0.07,
            '31421':0.07,
            '31501':0.08,
            '31502':0.08,
            '31503':0.08,
            '31510':0.07,
            '31512':0.07,
            '31513':0.08,
            '31515':0.08,
            '31516':0.07,
            '31518':0.08,
            '31519':0.07,
            '31520':0.07,
            '31521':0.07,
            '31522':0.07,
            '31523':0.07,
            '31524':0.07,
            '31525':0.07,
            '31527':0.07,
            '31532':0.08,
            '31533':0.07,
            '31534':0.07,
            '31535':0.07,
            '31537':0.07,
            '31539':0.08,
            '31542':0.07,
            '31543':0.07,
            '31544':0.08,
            '31545':0.08,
            '31546':0.08,
            '31547':0.07,
            '31548':0.07,
            '31549':0.08,
            '31550':0.08,
            '31551':0.07,
            '31552':0.08,
            '31553':0.07,
            '31554':0.07,
            '31555':0.08,
            '31556':0.07,
            '31557':0.07,
            '31558':0.07,
            '31560':0.08,
            '31561':0.07,
            '31562':0.07,
            '31563':0.08,
            '31564':0.08,
            '31565':0.07,
            '31566':0.07,
            '31567':0.07,
            '31568':0.07,
            '31569':0.07,
            '31598':0.08,
            '31599':0.08,
            '31601':0.07,
            '31602':0.07,
            '31603':0.07,
            '31604':0.07,
            '31605':0.07,
            '31606':0.07,
            '31620':0.07,
            '31622':0.07,
            '31623':0.07,
            '31624':0.07,
            '31625':0.07,
            '31626':0.07,
            '31627':0.07,
            '31629':0.07,
            '31630':0.07,
            '31631':0.07,
            '31632':0.07,
            '31634':0.07,
            '31635':0.07,
            '31636':0.07,
            '31637':0.07,
            '31638':0.07,
            '31639':0.07,
            '31641':0.07,
            '31642':0.07,
            '31643':0.07,
            '31645':0.07,
            '31647':0.07,
            '31648':0.07,
            '31649':0.07,
            '31650':0.07,
            '31698':0.07,
            '31699':0.07,
            '31701':0.07,
            '31702':0.07,
            '31703':0.07,
            '31704':0.07,
            '31705':0.07,
            '31706':0.07,
            '31707':0.07,
            '31708':0.07,
            '31709':0.08,
            '31711':0.08,
            '31712':0.08,
            '31714':0.07,
            '31716':0.07,
            '31719':0.08,
            '31720':0.07,
            '31721':0.07,
            '31722':0.08,
            '31727':0.07,
            '31730':0.07,
            '31733':0.07,
            '31735':0.08,
            '31738':0.07,
            '31739':0.07,
            '31743':0.08,
            '31744':0.08,
            '31747':0.08,
            '31749':0.07,
            '31750':0.07,
            '31753':0.08,
            '31756':0.08,
            '31757':0.07,
            '31758':0.07,
            '31763':0.07,
            '31764':0.08,
            '31765':0.07,
            '31768':0.08,
            '31769':0.07,
            '31771':0.08,
            '31772':0.07,
            '31773':0.07,
            '31774':0.07,
            '31775':0.08,
            '31776':0.08,
            '31778':0.07,
            '31779':0.07,
            '31780':0.08,
            '31781':0.07,
            '31782':0.07,
            '31783':0.07,
            '31784':0.07,
            '31787':0.07,
            '31788':0.08,
            '31789':0.07,
            '31790':0.07,
            '31791':0.07,
            '31792':0.07,
            '31793':0.07,
            '31794':0.07,
            '31795':0.07,
            '31796':0.07,
            '31798':0.07,
            '31799':0.07,
            '31801':0.08,
            '31803':0.08,
            '31804':0.08,
            '31805':0.08,
            '31806':0.08,
            '31807':0.08,
            '31808':0.08,
            '31810':0.08,
            '31811':0.08,
            '31812':0.08,
            '31814':0.08,
            '31815':0.08,
            '31816':0.07,
            '31820':0.08,
            '31821':0.08,
            '31822':0.08,
            '31823':0.08,
            '31824':0.08,
            '31825':0.08,
            '31826':0.08,
            '31827':0.08,
            '31829':0.08,
            '31830':0.07,
            '31831':0.08,
            '31832':0.08,
            '31833':0.07,
            '31836':0.08,
            '31901':0.08,
            '31902':0.08,
            '31903':0.08,
            '31904':0.08,
            '31905':0.08,
            '31906':0.08,
            '31907':0.08,
            '31908':0.08,
            '31909':0.08,
            '31914':0.08,
            '31917':0.08,
            '31993':0.08,
            '31995':0.08,
            '31997':0.08,
            '31998':0.08,
            '31999':0.08,
            '39813':0.07,
            '39815':0.08,
            '39817':0.08,
            '39818':0.08,
            '39819':0.08,
            '39823':0.07,
            '39824':0.08,
            '39825':0.08,
            '39826':0.07,
            '39827':0.07,
            '39828':0.07,
            '39829':0.07,
            '39832':0.07,
            '39834':0.08,
            '39836':0.08,
            '39837':0.07,
            '39840':0.08,
            '39841':0.07,
            '39842':0.07,
            '39845':0.07,
            '39846':0.07,
            '39851':0.08,
            '39852':0.08,
            '39854':0.08,
            '39859':0.07,
            '39861':0.07,
            '39862':0.07,
            '39866':0.07,
            '39867':0.08,
            '39870':0.07,
            '39877':0.07,
            '39885':0.07,
            '39886':0.08,
            '39897':0.07,
            '39901':0.08,

            '29001':0.08,
            '29002':0.08,
            '29003':0.08,
            '29006':0.07,
            '29009':0.08,
            '29010':0.08,
            '29014':0.08,
            '29015':0.07,
            '29016':0.08,
            '29018':0.07,
            '29020':0.08,
            '29021':0.08,
            '29030':0.07,
            '29031':0.07,
            '29032':0.08,
            '29033':0.07,
            '29036':0.07,
            '29037':0.07,
            '29038':0.07,
            '29039':0.07,
            '29040':0.08,
            '29041':0.08,
            '29042':0.08,
            '29044':0.08,
            '29045':0.08,
            '29046':0.08,
            '29047':0.07,
            '29048':0.07,
            '29051':0.08,
            '29052':0.08,
            '29053':0.07,
            '29054':0.07,
            '29055':0.08,
            '29056':0.08,
            '29058':0.08,
            '29059':0.07,
            '29061':0.08,
            '29062':0.08,
            '29063':0.08,
            '29065':0.07,
            '29067':0.08,
            '29069':0.08,
            '29070':0.07,
            '29071':0.07,
            '29072':0.07,
            '29073':0.07,
            '29074':0.08,
            '29075':0.07,
            '29078':0.08,
            '29079':0.08,
            '29080':0.08,
            '29081':0.08,
            '29082':0.08,
            '29101':0.08,
            '29102':0.08,
            '29104':0.08,
            '29105':0.08,
            '29107':0.07,
            '29108':0.07,
            '29111':0.08,
            '29112':0.07,
            '29113':0.07,
            '29114':0.08,
            '29115':0.07,
            '29116':0.07,
            '29117':0.07,
            '29118':0.07,
            '29122':0.07,
            '29123':0.07,
            '29125':0.08,
            '29126':0.07,
            '29127':0.07,
            '29128':0.08,
            '29129':0.08,
            '29130':0.07,
            '29132':0.07,
            '29133':0.07,
            '29135':0.07,
            '29137':0.08,
            '29138':0.07,
            '29142':0.07,
            '29143':0.08,
            '29145':0.07,
            '29146':0.07,
            '29147':0.08,
            '29148':0.08,
            '29150':0.08,
            '29151':0.08,
            '29152':0.08,
            '29153':0.08,
            '29154':0.08,
            '29160':0.07,
            '29161':0.08,
            '29162':0.08,
            '29163':0.07,
            '29164':0.08,
            '29166':0.07,
            '29168':0.08,
            '29169':0.07,
            '29170':0.07,
            '29171':0.07,
            '29172':0.07,
            '29175':0.08,
            '29177':0.08,
            '29178':0.07,
            '29180':0.07,
            '29201':0.08,
            '29202':0.08,
            '29203':0.08,
            '29204':0.08,
            '29205':0.08,
            '29206':0.08,
            '29207':0.08,
            '29208':0.08,
            '29209':0.08,
            '29210':0.08,
            '29211':0.08,
            '29212':0.07,
            '29214':0.08,
            '29215':0.08,
            '29216':0.08,
            '29217':0.08,
            '29218':0.08,
            '29219':0.08,
            '29220':0.08,
            '29221':0.08,
            '29222':0.08,
            '29223':0.08,
            '29224':0.08,
            '29225':0.08,
            '29226':0.08,
            '29227':0.08,
            '29228':0.07,
            '29229':0.08,
            '29230':0.08,
            '29240':0.08,
            '29250':0.08,
            '29260':0.08,
            '29290':0.08,
            '29292':0.08,
            '29301':0.07,
            '29302':0.07,
            '29303':0.07,
            '29304':0.07,
            '29305':0.07,
            '29306':0.07,
            '29307':0.07,
            '29316':0.07,
            '29319':0.07,
            '29320':0.07,
            '29321':0.07,
            '29322':0.07,
            '29323':0.07,
            '29324':0.07,
            '29325':0.07,
            '29329':0.07,
            '29330':0.07,
            '29331':0.07,
            '29332':0.07,
            '29333':0.07,
            '29334':0.07,
            '29335':0.07,
            '29336':0.07,
            '29338':0.07,
            '29340':0.08,
            '29341':0.08,
            '29342':0.08,
            '29346':0.07,
            '29348':0.07,
            '29349':0.07,
            '29351':0.07,
            '29353':0.07,
            '29355':0.07,
            '29356':0.06,
            '29360':0.07,
            '29364':0.07,
            '29365':0.07,
            '29368':0.07,
            '29369':0.07,
            '29370':0.07,
            '29372':0.07,
            '29373':0.07,
            '29374':0.07,
            '29375':0.07,
            '29376':0.07,
            '29377':0.07,
            '29378':0.07,
            '29379':0.07,
            '29384':0.07,
            '29385':0.07,
            '29386':0.07,
            '29388':0.07,
            '29395':0.07,
            '29401':0.09,
            '29402':0.09,
            '29403':0.09,
            '29404':0.09,
            '29405':0.09,
            '29406':0.09,
            '29407':0.09,
            '29409':0.09,
            '29410':0.08,
            '29412':0.09,
            '29413':0.09,
            '29414':0.09,
            '29415':0.09,
            '29416':0.09,
            '29417':0.09,
            '29418':0.09,
            '29419':0.09,
            '29420':0.07,
            '29422':0.09,
            '29423':0.09,
            '29424':0.09,
            '29425':0.09,
            '29426':0.09,
            '29429':0.09,
            '29431':0.08,
            '29432':0.07,
            '29433':0.08,
            '29434':0.08,
            '29435':0.08,
            '29436':0.08,
            '29437':0.07,
            '29438':0.09,
            '29439':0.09,
            '29440':0.07,
            '29442':0.07,
            '29445':0.08,
            '29446':0.08,
            '29447':0.07,
            '29448':0.07,
            '29449':0.09,
            '29450':0.08,
            '29451':0.09,
            '29452':0.08,
            '29453':0.08,
            '29455':0.09,
            '29456':0.07,
            '29457':0.09,
            '29458':0.09,
            '29461':0.08,
            '29464':0.09,
            '29465':0.09,
            '29466':0.09,
            '29468':0.08,
            '29469':0.08,
            '29470':0.09,
            '29471':0.07,
            '29472':0.07,
            '29474':0.08,
            '29475':0.08,
            '29476':0.08,
            '29477':0.07,
            '29479':0.08,
            '29481':0.08,
            '29482':0.09,
            '29483':0.07,
            '29484':0.07,
            '29485':0.07,
            '29486':0.08,
            '29487':0.09,
            '29488':0.08,
            '29492':0.08,
            '29493':0.08,
            '29501':0.08,
            '29502':0.08,
            '29503':0.08,
            '29504':0.08,
            '29505':0.08,
            '29506':0.08,
            '29510':0.08,
            '29511':0.08,
            '29512':0.08,
            '29516':0.08,
            '29518':0.08,
            '29519':0.08,
            '29520':0.08,
            '29525':0.08,
            '29526':0.08,
            '29527':0.08,
            '29528':0.08,
            '29530':0.08,
            '29532':0.08,
            '29536':0.08,
            '29540':0.08,
            '29541':0.08,
            '29543':0.08,
            '29544':0.08,
            '29545':0.08,
            '29546':0.08,
            '29547':0.08,
            '29550':0.08,
            '29551':0.08,
            '29554':0.07,
            '29555':0.08,
            '29556':0.08,
            '29560':0.08,
            '29563':0.08,
            '29564':0.08,
            '29565':0.08,
            '29566':0.08,
            '29567':0.08,
            '29568':0.08,
            '29569':0.08,
            '29570':0.08,
            '29571':0.08,
            '29572':0.09,
            '29574':0.08,
            '29575':0.08,
            '29576':0.08,
            '29577':0.09,
            '29578':0.09,
            '29579':0.08,
            '29580':0.08,
            '29581':0.08,
            '29582':0.08,
            '29583':0.08,
            '29584':0.08,
            '29585':0.07,
            '29587':0.08,
            '29588':0.08,
            '29589':0.08,
            '29590':0.08,
            '29591':0.08,
            '29592':0.08,
            '29593':0.08,
            '29594':0.08,
            '29596':0.08,
            '29597':0.08,
            '29598':0.08,
            '29601':0.06,
            '29602':0.06,
            '29603':0.06,
            '29604':0.06,
            '29605':0.06,
            '29606':0.06,
            '29607':0.06,
            '29608':0.06,
            '29609':0.06,
            '29610':0.06,
            '29611':0.06,
            '29612':0.06,
            '29613':0.06,
            '29614':0.06,
            '29615':0.06,
            '29616':0.06,
            '29617':0.06,
            '29620':0.07,
            '29621':0.07,
            '29622':0.07,
            '29623':0.07,
            '29624':0.07,
            '29625':0.07,
            '29626':0.07,
            '29627':0.07,
            '29628':0.07,
            '29630':0.07,
            '29631':0.07,
            '29632':0.07,
            '29633':0.07,
            '29634':0.07,
            '29635':0.06,
            '29636':0.06,
            '29638':0.07,
            '29639':0.07,
            '29640':0.07,
            '29641':0.07,
            '29642':0.07,
            '29643':0.06,
            '29644':0.07,
            '29645':0.07,
            '29646':0.07,
            '29647':0.07,
            '29648':0.07,
            '29649':0.07,
            '29650':0.06,
            '29651':0.06,
            '29652':0.06,
            '29653':0.07,
            '29654':0.07,
            '29655':0.07,
            '29656':0.07,
            '29657':0.07,
            '29658':0.06,
            '29659':0.07,
            '29661':0.06,
            '29662':0.06,
            '29664':0.06,
            '29665':0.06,
            '29666':0.07,
            '29667':0.07,
            '29669':0.06,
            '29670':0.07,
            '29671':0.07,
            '29672':0.06,
            '29673':0.07,
            '29675':0.06,
            '29676':0.06,
            '29677':0.07,
            '29678':0.06,
            '29679':0.06,
            '29680':0.06,
            '29681':0.06,
            '29682':0.07,
            '29683':0.06,
            '29684':0.07,
            '29685':0.07,
            '29686':0.06,
            '29687':0.06,
            '29688':0.06,
            '29689':0.07,
            '29690':0.06,
            '29691':0.06,
            '29692':0.07,
            '29693':0.06,
            '29696':0.06,
            '29697':0.07,
            '29702':0.08,
            '29703':0.07,
            '29704':0.07,
            '29706':0.08,
            '29707':0.08,
            '29708':0.07,
            '29709':0.08,
            '29710':0.07,
            '29712':0.08,
            '29714':0.08,
            '29715':0.07,
            '29716':0.07,
            '29717':0.07,
            '29718':0.08,
            '29720':0.08,
            '29721':0.08,
            '29722':0.08,
            '29724':0.08,
            '29726':0.07,
            '29727':0.08,
            '29728':0.08,
            '29729':0.08,
            '29730':0.07,
            '29731':0.07,
            '29732':0.07,
            '29733':0.07,
            '29734':0.07,
            '29741':0.08,
            '29742':0.07,
            '29743':0.07,
            '29744':0.08,
            '29745':0.07,
            '29801':0.08,
            '29802':0.08,
            '29803':0.08,
            '29804':0.08,
            '29805':0.08,
            '29808':0.08,
            '29809':0.08,
            '29810':0.08,
            '29812':0.08,
            '29813':0.08,
            '29816':0.08,
            '29817':0.08,
            '29819':0.07,
            '29821':0.08,
            '29822':0.08,
            '29824':0.07,
            '29826':0.08,
            '29827':0.08,
            '29828':0.08,
            '29829':0.08,
            '29831':0.08,
            '29832':0.07,
            '29834':0.08,
            '29835':0.08,
            '29836':0.08,
            '29838':0.08,
            '29839':0.08,
            '29840':0.08,
            '29841':0.08,
            '29842':0.08,
            '29843':0.08,
            '29844':0.08,
            '29845':0.08,
            '29846':0.08,
            '29847':0.07,
            '29848':0.08,
            '29849':0.08,
            '29850':0.08,
            '29851':0.08,
            '29853':0.08,
            '29856':0.08,
            '29860':0.07,
            '29861':0.08,
            '29899':0.08,
            '29901':0.06,
            '29902':0.06,
            '29903':0.06,
            '29904':0.06,
            '29905':0.06,
            '29906':0.06,
            '29907':0.06,
            '29909':0.06,
            '29910':0.06,
            '29911':0.08,
            '29912':0.09,
            '29913':0.08,
            '29914':0.06,
            '29915':0.06,
            '29916':0.08,
            '29918':0.08,
            '29920':0.06,
            '29921':0.08,
            '29922':0.08,
            '29923':0.08,
            '29924':0.08,
            '29925':0.06,
            '29926':0.06,
            '29927':0.09,
            '29928':0.06,
            '29929':0.08,
            '29931':0.06,
            '29932':0.08,
            '29933':0.08,
            '29934':0.09,
            '29935':0.06,
            '29936':0.09,
            '29938':0.06,
            '29939':0.08,
            '29940':0.06,
            '29941':0.06,
            '29943':0.09,
            '29944':0.08,
            '29945':0.08}
        tax = zipdict[zip]
        tax = tax*100
        tax = round(tax,2)

            
            
        self.taxRate.setText('Tax Rate: '+str(tax)+'%') 
        
        quotedPrice = re.sub('[!@#$,]', '', quotedPrice)
        quotedPrice = float(quotedPrice)
        
        taxCal = tax/100
        totalWithTax= (quotedPrice*taxCal)+quotedPrice
        
        if self.taxFree.isChecked():
            totalWithTax = quotedPrice
        
            
        self.inputTotalPrice.setText(str(totalWithTax))
        
        
    def quoteForm(self):
        
        try:
        
            jobNumber = self.inputJobNumber.text() 
            
            o = win32com.client.Dispatch("Excel.Application")

            o.Visible = False

            wb_path = os.getcwd()+'\\' + jobNumber + '.xlsx'

            wb = o.Workbooks.Open(wb_path)



            ws_index_list = [2] #say you want to print these sheets

            path_to_pdf = os.getcwd()+'\\' + jobNumber + ' Quote.pdf'

            print_area = 'A1:K42'
            
            



            for index in ws_index_list:

                #off-by-one so the user can start numbering the worksheets at 1

                ws = wb.Worksheets[index - 1]

                ws.PageSetup.Zoom = False

                ws.PageSetup.FitToPagesTall = 1

                ws.PageSetup.FitToPagesWide = 1

                ws.PageSetup.PrintArea = print_area



            wb.WorkSheets(ws_index_list).Select()

            wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
            wb.Close(True)
            pyautogui.alert('Quote Form Created')
            if not os.path.exists(r'O:\Jobs\\'+jobNumber+'\Quote'):
                os.mkdir(r'O:\Jobs\\'+jobNumber+'\Quote')
            
            file = jobNumber+' Quote.pdf'
            shutil.move(os.getcwd()+'\\'+file,'O:\Jobs\\'+jobNumber+'\Quote\\'+file)


        
            pyautogui.alert('did it work?')
        except:
            pyautogui.alert('Unable to creatE pdf Have you Consolidated this job yet?')           



        
    def jobManagment(self):
        
        jobNameEst = self.jobNameEst.text()
        priceEst = self.lPrice.text()
        bfEst = lTotalBf = self.lTotalBf.text()
        
        runNameEst = self.runNameEst.text()
        

        
         
        if not os.path.exists( jobNameEst):
            os.mkdir(jobNameEst)        
        
        file = jobNameEst + runNameEst
        jobEstInfo = {'priceEst': priceEst, 'bfEst' : bfEst} 
        pickle.dump( jobEstInfo, open(file, "wb"))
        
        shutil.move(os.getcwd()+'\\'+file,os.getcwd()+'\\'+jobNameEst+'\\'+file)
        
        rowPosition = self.tableJobEst.rowCount()
        self.tableJobEst.insertRow(rowPosition)
        self.tableJobEst.setItem(rowPosition , 0, PySide2.QtWidgets.QTableWidgetItem(jobNameEst+ ' '+ runNameEst))
        self.tableJobEst.setItem(rowPosition , 1, PySide2.QtWidgets.QTableWidgetItem(bfEst))
        self.tableJobEst.setItem(rowPosition , 2, PySide2.QtWidgets.QTableWidgetItem(priceEst))
        
    def jobTotal(self):
        val = sum([float(item.text()) for item in self.tableJobEst.selectedItems()])
        table = PySide2.QtWidgets.QTableWidgetItem()
        #table.setText(str(val))
        val = round(val,2)
        #row = self.tableJobEst.currentRow()
        #self.tableJobEst.setItem(row, 2, table)
        rowPosition = self.tableJobEst.rowCount()
        self.tableJobEst.insertRow(rowPosition)
        self.tableJobEst.setItem(rowPosition , 0, PySide2.QtWidgets.QTableWidgetItem('0'))
        self.tableJobEst.setItem(rowPosition , 1, PySide2.QtWidgets.QTableWidgetItem('0'))
        self.tableJobEst.setItem(rowPosition , 2, PySide2.QtWidgets.QTableWidgetItem('$' + str(val)))   
        
    def findJob(self):
        self.tableJobEst.setRowCount(0)
        jobNameEst = self.jobNameEst.text()
        if os.path.exists( jobNameEst):
            path = jobNameEst

            files = []
                # r=root, d=directories, f = files
            for r, d, f in os.walk(path):
                for file in f:
                    files.append(os.path.join(file))

            for f in files:
                jobInfo = pickle.load( open( path+'\\'+f, "rb" ))
                priceEst = jobInfo.get('priceEst','')
                bfEst = jobInfo.get('bfEst','')
                
                rowPosition = self.tableJobEst.rowCount()
                self.tableJobEst.insertRow(rowPosition)
                self.tableJobEst.setItem(rowPosition , 0, PySide2.QtWidgets.QTableWidgetItem(f))
                self.tableJobEst.setItem(rowPosition , 1, PySide2.QtWidgets.QTableWidgetItem(bfEst))
                self.tableJobEst.setItem(rowPosition , 2, PySide2.QtWidgets.QTableWidgetItem(priceEst))
        
            
    def deleteSel(self):
        jobNameEst = self.jobNameEst.text()
        path = jobNameEst
        selected = self.tableJobEst.currentRow()
        fileName = self.tableJobEst.item(selected,0)
        self.ID = fileName.text()
        os.remove(path+'\\'+self.ID)
        
        self.tableJobEst.removeRow(selected)
        
    def bfEstFloor(self): 
        depth= self.depthFloor.currentText()
        span = self.spanFloor.value()
        trussNumber = self.trussNumberFloor.text()
        
        if depth == '12" - 16"':
            price = span * 2.75
        
        if depth == '18" - 24"':
            price = span * 3.20
        
        
        
        self.lPrice.setText(str(price)) 
        
        
        
 
if __name__ == '__main__':
    app = QApplication(sys.argv)
    form = Form('mainwindow.ui')
    sys.exit(app.exec_())